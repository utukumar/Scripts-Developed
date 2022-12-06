#importing the necessary libraries required
import encodings
from mimetypes import encodings_map
import sys
from pathlib import Path
import pandas as pd
import os
import xlrd
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import filedialog
from tkinter.filedialog import askopenfile
from tags import *

#creating pop-up open window for selecting the downloaded excel sheet
root = tk.Tk()
root.withdraw()
root.attributes("-topmost", True)
file_path = filedialog.askopenfile(mode = 'r')
#assigning the absolute path to the path variable
path = os.path.abspath(file_path.name)

#reading excel sheet
excel_worksheet = xlrd.open_workbook(path)
excel_sheet = excel_worksheet.sheet_by_index(0)

#reading the excel sheet as a pandas data frame
df = pd.read_excel(path, "Data")

#path for the MBR Format excel sheet where the data needs to be written
path_write_xl = r'C:\Users\utukumar\Documents\Scripts\dist\Prime\Report Generation\MBR format.xlsx'

#calculating various tag count for different verticals
def calculating_issues(row_length, column_tag, vertical_tag):
    count = 0
    for i in range(0,row_length):
        if (column_tag in df.iloc[i,-1]) and (vertical_tag in df.iloc[i,-1]):
            count +=1
    return count

#calculating blocker issues for different verticals
def calculating_blocker_issues(row_length, vertical_tag):
    count = 0
    for i in range(0,row_length):
        if ((blocker_issue_tag_one in df.iloc[i,-1]) or (blocker_issue_tag_two in df.iloc[i,-1]) or (blocker_issue_tag_three in df.iloc[i,-1])) and (vertical_tag in df.iloc[i,-1]):
            count += 1
    return count

#calculating open and resolved blocker issues for different verticals
def calcualting_open_and_resolved_blocker_issues(row_length, vertical_tag):
    
    #the function return two values first is blocker_open count and second is blocker_resolved count
    blocker_open, blocker_resolved = 0, 0
    for i in range(0,row_length):
        if (((blocker_issue_tag_one in df.iloc[i,-1]) or (blocker_issue_tag_two in df.iloc[i,-1]) or (blocker_issue_tag_three in df.iloc[i,-1])) and (vertical_tag in df.iloc[i,-1])) and ('Open' in df.iloc[i,3]):
            blocker_open += 1
        elif (((blocker_issue_tag_one in df.iloc[i,-1]) or (blocker_issue_tag_two in df.iloc[i,-1]) or (blocker_issue_tag_three in df.iloc[i,-1])) and (vertical_tag in df.iloc[i,-1])) and ('Resolved' in df.iloc[i,3] or 'Closed' in df.iloc[i,3]):
            blocker_resolved += 1
    
    #returning two values first -- blocker_open, second -- blocker_resolved
    return blocker_open, blocker_resolved            
                                 
#calculating total DA_Acquisition issues
for i in range(0,len(df)):
    if acquistion_tag in df.iloc[i,-1]:
        acquistion_issues+=1
print("Total Acquistion issues : ",acquistion_issues)
        
#calculating resolved issues for DA_Acquisition issue
for i in range(0, len(df)):
    if ('Closed' in df.iloc[i,3] or 'Resolved' in df.iloc[i,3]) and (acquistion_tag in df.iloc[i,-1]):
        resolved_acquistion_issue += 1
print("Resolved Acquistion Issue : ",resolved_acquistion_issue)

#calculating open issues for DA_Acquisition issue
for i in range(1,len(df)):
    if 'Open' in df.iloc[i,3] and acquistion_tag in df.iloc[i,-1]:
        open_acquistion_issue += 1
print("Unresolved Acquistion Issue : ",open_acquistion_issue)

#calculating fixed issues for DA_Acquisition vertical
fixed_acquistion_issue = calculating_issues(len(df), fixed_tag, acquistion_tag)
print("Fixed Acquistion Issues : ",fixed_acquistion_issue)

#calculating cannot reproduce issues for DA_Acquistion Vertical
cantreproduce_acquistion_issue = calculating_issues(len(df), cantreproduce_tag, acquistion_tag)
print("Cannot Reproduce Acquistion Issues : ",cantreproduce_acquistion_issue)

#calculating duplicate issues for DA_Acquistion Vertical
duplicate_acquistion_issue = calculating_issues(len(df), duplicate_tag, acquistion_tag)
print("Duplicate Acquistion Issues : ",duplicate_acquistion_issue)

#calculating triaged issues for DA_Acquistion Vertical
triaged_acquistion_issue = calculating_issues(len(df), triaged_tag, acquistion_tag)
print("Triaged Acquistion Issues : ",triaged_acquistion_issue)

#calcualting deferred issues for DA_Acquistion Vertical
deferred_acquistion_issue = calculating_issues(len(df), deferred_tag, acquistion_tag)
print("Deferred Acquisition Issue : ",deferred_acquistion_issue)

#calculating rca_dispute issue for DA_Acquistion vertical
rcadispute_acuistion_issue = calculating_issues(len(df), rcadispute_tag, acquistion_tag)
print("RCA_Dispute Acquistion Issue : ",rcadispute_acquistion_issue)

#calculating tc issue for DA_Acquistion vertical
tc_acquistion_issue = calculating_issues(len(df), tc_tag, acquistion_tag)
print("TC Acquistion Issue : ",tc_acquistion_issue)

#calculating adhoc issue for DA_Acquistion vertical
adhoc_acquistion_issue = calculating_issues(len(df), adhoc_tag, acquistion_tag)
print("Adhoc Acquistion Issue : ",adhoc_acquistion_issue)

#calculating moved_to_tt_issues for DA_Acquistion vertical
moved_to_tt_acquistion_issue = calculating_issues(len(df), moved_to_tt_tag, acquistion_tag)
print("Moved to TT Acquistion Issue : ", moved_to_tt_acquistion_issue)

#calculating not_fixed_valid for DA_Acquistion vertical
notfixed_valid_acquistion_issue = calculating_issues(len(df), notfixed_valid_tag, acquistion_tag)
print("Not Fixed Valid Acquistion Issue : ", notfixed_valid_acquistion_issue)

#calculating pdfixed for DA_Acquistion vertical
pdfixed_acquistion_issue = calculating_issues(len(df), pdfixed_tag, acquistion_tag)
print("PD_Fixed Acquistion Issue : ", pdfixed_acquistion_issue)

#calculating postlaunch_fixed for DA_Acquistion vertical
postlaunch_fix_acquistion_issue = calculating_issues(len(df), postlaunch_fix_tag, acquistion_tag)
print("Post Launch Fix Acquistion Issue : ", postlaunch_fix_acquistion_issue)

#calculating deployment issue for DA_Acquistion vertical
deployment_acquistion_issue = calculating_issues(len(df), deployment_tag, acquistion_tag)
print("Deployment Acquistion Issue : ", deployment_acquistion_issue)

#calculating bydesign_valid for DA_Acquistion vertical
bydesign_valid_acquistion_issue = calculating_issues(len(df), bydesign_valid_tag, acquistion_tag)
print("By Design Valid Acquistion Issue : ", bydesign_valid_acquistion_issue)

#calculating bydesign_tc for DA_Acquistion vertical
bydesign_tc_acquistion_issue = calculating_issues(len(df), bydesign_tc_tag, acquistion_tag)
print("ByDesign TC Acquistion Issue : ", bydesign_tc_acquistion_issue)

#calculating business accepted defect for DA_Acquistion vertical
business_accepted_defect_acquistion_issue = calculating_issues(len(df), business_accepted_defect_tag, acquistion_tag)
print("Business Accepted Defect for Acquistion : ", business_accepted_defect_acquistion_issue)

#calculating known issue for DA_Acquistion vertical
knownissue_acquistion_issue = calculating_issues(len(df), knownissue_tag, acquistion_tag)
print("Known Issue for Acquistion Vertical : ", knownissue_acquistion_issue)

#calulating blocker issue for DA_Acquistion vertical
blocker_acquistion_issue = calculating_blocker_issues(len(df), acquistion_tag)
print("DA_Acquistion Blocker Issues : ",blocker_acquistion_issue)

#calculating open and resolved blocker bugs for DA_Acquistion vertical
blocker_acquistion_open_issues, blocker_acquistion_resolved_issues = calcualting_open_and_resolved_blocker_issues(len(df), acquistion_tag)
print("DA_Acquistion Open Blocker Issues : ", blocker_acquistion_open_issues)
print("DA_Acquistion Resolved Blocker Issues : ", blocker_acquistion_resolved_issues)
        
        

#calculating valid issues for DA_Acquistion vertical
valid_acquistion_issue = (fixed_acquistion_issue+notfixed_valid_acquistion_issue+pdfixed_acquistion_issue+postlaunch_fix_acquistion_issue+ 
                          deferred_acquistion_issue+deployment_acquistion_issue+rcadispute_acquistion_issue+bydesign_valid_acquistion_issue+
                          bydesign_tc_acquistion_issue+business_accepted_defect_acquistion_issue)
print('Valid defects in DA_Acquistion vertical : ',valid_acquistion_issue)

#valid defect ration for DA_Acquistion Vertical
if resolved_acquistion_issue == 0:
    valid_defect_ratio_acquistion = 0
else:
    valid_defect_ratio_acquistion = int(valid_acquistion_issue*100/resolved_acquistion_issue)
print(str('Valid defect ration in DA_Acquistion : ')+str(valid_defect_ratio_acquistion)+'%') 

#defect fix ratio for DA_Acquistion Vertical
if resolved_acquistion_issue == 0:
    defect_fix_ratio_acquistion = 0
else:
    defect_fix_ratio_acquistion = int(fixed_acquistion_issue*100/resolved_acquistion_issue)
print(str('Defect Fix Ratio in DA_Acquistion : ')+str(defect_fix_ratio_acquistion)+'%')

#writing the values to MBR Format excel sheet for DA_Acquistion
wb = load_workbook(path_write_xl)
#selceting the active sheet from the excel file, active sheet is generally the sheet that we see when the excel file is opened
ws = wb.active
ws['B10'].value = acquistion_issues
ws['B11'].value = tc_acquistion_issue
ws['B12'].value = adhoc_acquistion_issue
ws['B13'].value = valid_acquistion_issue
ws['B14'].value = (bydesign_tc_acquistion_issue+bydesign_valid_acquistion_issue)
ws['B15'].value = (rcadispute_acuistion_issue+deployment_acquistion_issue)
ws['B16'].value = moved_to_tt_acquistion_issue
ws['B17'].value = (business_accepted_defect_acquistion_issue+knownissue_acquistion_issue)
ws['B18'].value = deployment_acquistion_issue
ws['B20'].value = cantreproduce_acquistion_issue
ws['B21'].value = duplicate_acquistion_issue
ws['B23'].value = fixed_acquistion_issue + pdfixed_acquistion_issue + postlaunch_fix_acquistion_issue
ws['B24'].value = resolved_acquistion_issue
ws['B25'].value = open_acquistion_issue
ws['B26'].value = blocker_acquistion_issue
ws['B27'].value = blocker_acquistion_open_issues
ws['B28'].value = blocker_acquistion_resolved_issues
ws['B34'].value = triaged_acquistion_issue
ws['B35'].value = valid_defect_ratio_acquistion
ws['B36'].value = defect_fix_ratio_acquistion

#closing the opened excel file after writing in it
wb.save(path_write_xl)

#calculating total DA_WLP issues
for i in range(0,len(df)):
    if wlp_tag in df.iloc[i,-1]:
        wlp_issues+=1
print('\n')
print("Total WLP issues : ",wlp_issues)
        
#calculating resolved issues for DA_WLP issue
for i in range(0, len(df)):
    if ('Closed' in df.iloc[i,3] or 'Resolved' in df.iloc[i,3]) and (wlp_tag in df.iloc[i,-1]):
        resolved_wlp_issue += 1
print("Resolved WLP Issue : ",resolved_wlp_issue)

#calculating open issues for DA_WLP issue
for i in range(1,len(df)):
    if 'Open' in df.iloc[i,3] and wlp_tag in df.iloc[i,-1]:
        open_wlp_issue += 1
print("Unresolved WLP Issue : ",open_wlp_issue)

#calculating fixed issues for DA_WLP vertical
fixed_wlp_issue = calculating_issues(len(df), fixed_tag, wlp_tag)
print("Fixed WLP Issues : ",fixed_wlp_issue)

#calculating cannot reproduce issues for DA_WLP Vertical
cantreproduce_wlp_issue = calculating_issues(len(df), cantreproduce_tag, wlp_tag)
print("Cannot Reproduce WLP Issues : ",cantreproduce_wlp_issue)

#calculating duplicate issues for DA_WLP Vertical
duplicate_wlp_issue = calculating_issues(len(df), duplicate_tag, wlp_tag)
print("Duplicate WLP Issues : ",duplicate_wlp_issue)

#calculating triaged issues for DA_WLP Vertical
triaged_wlp_issue = calculating_issues(len(df), triaged_tag, wlp_tag)
print("Triaged WLP Issues : ",triaged_wlp_issue)

#calcualting deferred issues for DA_WLP Vertical
deferred_wlp_issue = calculating_issues(len(df), deferred_tag, wlp_tag)
print("Deferred WLP Issue : ",deferred_wlp_issue)

#calculating rca_dispute issue for DA_WLP vertical
rcadispute_wlp_issue = calculating_issues(len(df), rcadispute_tag, wlp_tag)
print("RCA_Dispute WLP Issue : ",rcadispute_wlp_issue)

#calculating tc issue for DA_WLP vertical
tc_wlp_issue = calculating_issues(len(df), tc_tag, wlp_tag)
print("TC WLP Issue : ",tc_wlp_issue)

#calculating adhoc issue for DA_WLP vertical
adhoc_wlp_issue = calculating_issues(len(df), adhoc_tag, wlp_tag)
print("Adhoc WLP Issue : ",adhoc_wlp_issue)

#calculating moved_to_tt_issues for DA_WLP vertical
moved_to_tt_wlp_issue = calculating_issues(len(df), moved_to_tt_tag, wlp_tag)
print("Moved to TT WLP Issue : ", moved_to_tt_wlp_issue)

#calculating not_fixed_valid for DA_WLP vertical
notfixed_valid_wlp_issue = calculating_issues(len(df), notfixed_valid_tag, wlp_tag)
print("Not Fixed Valid WLP Issue : ", notfixed_valid_wlp_issue)

#calculating pdfixed for DA_WLP vertical
pdfixed_wlp_issue = calculating_issues(len(df), pdfixed_tag, wlp_tag)
print("PD_Fixed WLP Issue : ", pdfixed_wlp_issue)

#calculating postlaunch_fixed for DA_WLP vertical
postlaunch_fix_wlp_issue = calculating_issues(len(df), postlaunch_fix_tag, wlp_tag)
print("Post Launch Fix WLP Issue : ", postlaunch_fix_wlp_issue)

#calculating deployment issue for DA_WLP vertical
deployment_wlp_issue = calculating_issues(len(df), deployment_tag, wlp_tag)
print("Deployment WLP Issue : ", deployment_wlp_issue)

#calculating bydesign_valid for DA_WLP vertical
bydesign_valid_wlp_issue = calculating_issues(len(df), bydesign_valid_tag, wlp_tag)
print("By Design Valid WLP Issue : ", bydesign_valid_wlp_issue)

#calculating bydesign_tc for DA_WLP vertical
bydesign_tc_wlp_issue = calculating_issues(len(df), bydesign_tc_tag, wlp_tag)
print("ByDesign TC WLP Issue : ", bydesign_tc_wlp_issue)

#calculating business accepted defect for DA_WLP vertical
business_accepted_defect_wlp_issue = calculating_issues(len(df), business_accepted_defect_tag, wlp_tag)
print("Business Accepted Defect for WLP : ", business_accepted_defect_wlp_issue)

#calculating known issue for DA_WLP vertical
knownissue_wlp_issue = calculating_issues(len(df), knownissue_tag, wlp_tag)
print("Known Issue for  WLP Vertical : ", knownissue_wlp_issue)

#calulating blocker issue for DA_WLP vertical
blocker_wlp_tag = calculating_blocker_issues(len(df), wlp_tag)
print("DA_WLP Blocker Issues : ",blocker_wlp_tag)

#calculating open and resolved blocker bugs for DA_WLP vertical
blocker_wlp_tag_open_issues, blocker_wlp_tag_resolved_issues = calcualting_open_and_resolved_blocker_issues(len(df), wlp_tag)
print("DA_WLP Open Blocker Issues : ", blocker_wlp_tag_open_issues)
print("DA_WLP Resolved Blocker Issues : ", blocker_wlp_tag_resolved_issues)

#calculating valid issues for DA_WLP vertical
valid_wlp_issue = (fixed_wlp_issue+notfixed_valid_wlp_issue+pdfixed_wlp_issue+postlaunch_fix_wlp_issue+ 
                    deferred_wlp_issue+deployment_wlp_issue+rcadispute_wlp_issue+bydesign_valid_wlp_issue+
                    bydesign_tc_wlp_issue+business_accepted_defect_wlp_issue)
print('Valid defects in DA_WLP vertical : ',valid_wlp_issue)

#valid defect ration for DA_WLP Vertical
if resolved_wlp_issue == 0:
    valid_defect_ratio_wlp = 0
else:
    valid_defect_ratio_wlp = int(valid_wlp_issue*100/resolved_wlp_issue)
print(str('Valid defect ration in DA_WLP : ')+str(valid_defect_ratio_wlp)+'%')

#defect fix ratio for DA_WLP Vertical
if resolved_wlp_issue == 0:
    defect_fix_ratio_wlp = 0
else:
    defect_fix_ratio_wlp = int(fixed_wlp_issue*100/resolved_wlp_issue)
print(str('Defect Fix Ratio for DA_WLP : ')+str(defect_fix_ratio_wlp)+'%')

#writing the values to MBR Format excel sheet for DA_WLP Vertical
wb = load_workbook(path_write_xl)
#selceting the active sheet from the excel file, active sheet is generally the sheet that we see when the excel file is opened
ws = wb.active
ws['C10'].value = wlp_issues
ws['C11'].value = tc_wlp_issue
ws['C12'].value = adhoc_wlp_issue
ws['C13'].value = valid_wlp_issue
ws['C14'].value = bydesign_tc_wlp_issue+bydesign_valid_wlp_issue
ws['C15'].value = rcadispute_acuistion_issue+deployment_wlp_issue
ws['C16'].value = moved_to_tt_wlp_issue
ws['C17'].value = business_accepted_defect_wlp_issue+knownissue_wlp_issue
ws['C18'].value = deployment_wlp_issue
ws['C20'].value = cantreproduce_wlp_issue
ws['C21'].value = duplicate_wlp_issue
ws['C23'].value = fixed_wlp_issue + pdfixed_wlp_issue + postlaunch_fix_wlp_issue
ws['C24'].value = resolved_wlp_issue
ws['C25'].value = open_wlp_issue
ws['C26'].value = blocker_wlp_tag
ws['C27'].value = blocker_wlp_tag_open_issues
ws['C28'].value = blocker_wlp_tag_resolved_issues
ws['C34'].value = triaged_wlp_issue
ws['C35'].value = valid_defect_ratio_wlp
ws['C36'].value = defect_fix_ratio_wlp
#closing the opened excel file after writing in it
wb.save(path_write_xl)

#calculating total DA_Retention_ScPy issues
for i in range(0,len(df)):
    if retention_scpy_tag in df.iloc[i,-1]:
        retention_scpy_issues += 1
print('\n')
print("Total DA_Retention_ScPy Issues : ",retention_scpy_issues)

#calculating realoved issues for DA_Retention_ScPy
for i in range(0,len(df)):
    if (('Closed' in df.iloc[i,3] or 'Resolved' in df.iloc[i,3]) and retention_scpy_tag in df.iloc[i,-1]):
        resolved_retention_scpy_issue += 1
print('Resolves issues for DA_Retention_ScPy : ',resolved_retention_scpy_issue)

#calculating open issues for DA_Retention_ScPy issue
for i in range(1,len(df)):
    if 'Open' in df.iloc[i,3] and retention_scpy_tag in df.iloc[i,-1]:
        open_retention_scpy_issue += 1
print("Unresolved DA_Retention_ScPy Issue : ",open_retention_scpy_issue)

#calculating fixed issues for DA_Retention_ScPy vertical
fixed_retention_scpy_issue = calculating_issues(len(df), fixed_tag, retention_scpy_tag)
print("Fixed DA_Retention_ScPy Issues : ",fixed_retention_scpy_issue)

#calculating cannot reproduce issues for DA_Retention_ScPy Vertical
cantreproduce_retention_scpy_issue = calculating_issues(len(df), cantreproduce_tag, retention_scpy_tag)
print("Cannot Reproduce DA_Retention_ScPy Issues : ",cantreproduce_retention_scpy_issue)

#calculating duplicate issues for DA_Retention_ScPy Vertical
duplicate_retention_scpy_issue = calculating_issues(len(df), duplicate_tag, retention_scpy_tag)
print("Duplicate DA_Retention_ScPy Issues : ",duplicate_retention_scpy_issue)

#calculating triaged issues for DA_Retention_ScPy Vertical
triaged_retention_scpy_issue = calculating_issues(len(df), triaged_tag, retention_scpy_tag)
print("Triaged DA_Retention_ScPy Issues : ",triaged_retention_scpy_issue)

#calcualting deferred issues for DA_Retention_ScPy Vertical
deferred_retention_scpy_issue = calculating_issues(len(df), deferred_tag, retention_scpy_tag)
print("Deferred DA_Retention_ScPy Issue : ",deferred_retention_scpy_issue)

#calculating rca_dispute issue for DA_Retention_ScPy vertical
rcadispute_retention_scpy_issue = calculating_issues(len(df), rcadispute_tag, retention_scpy_tag)
print("RCA_Dispute DA_Retention_ScPy Issue : ",rcadispute_retention_scpy_issue)

#calculating tc issue for DA_Retention_ScPy vertical
tc_retention_scpy_issue = calculating_issues(len(df), tc_tag, retention_scpy_tag)
print("TC DA_Retention_ScPy Issue : ",tc_retention_scpy_issue)

#calculating adhoc issue for DA_Retention_ScPy vertical
adhoc_retention_scpy_issue = calculating_issues(len(df), adhoc_tag, retention_scpy_tag)
print("Adhoc DA_Retention_ScPy Issue : ",adhoc_retention_scpy_issue)

#calculating moved_to_tt_issues for DA_Retention_ScPy vertical
moved_to_tt_retention_scpy_issue = calculating_issues(len(df), moved_to_tt_tag, retention_scpy_tag)
print("Moved to TT for DA_Retention_ScPy: ", moved_to_tt_retention_scpy_issue)

#calculating not_fixed_valid for DA_Retention_ScPy vertical
notfixed_valid_retention_scpy_issue = calculating_issues(len(df), notfixed_valid_tag, retention_scpy_tag)
print("Not Fixed Valid DA_Retention_ScPy Issue : ", notfixed_valid_retention_scpy_issue)

#calculating pdfixed for DA_Retention_ScPy vertical
pdfixed_retention_scpy_issue = calculating_issues(len(df), pdfixed_tag, retention_scpy_tag)
print("PD_Fixed DA_Retention_ScPy Issue : ", pdfixed_retention_scpy_issue)

#calculating postlaunch_fixed for DA_Retention_ScPy vertical
postlaunch_fix_retention_scpy_issue = calculating_issues(len(df), postlaunch_fix_tag, retention_scpy_tag)
print("Post Launch Fix DA_Retention_ScPy Issue : ", postlaunch_fix_retention_scpy_issue)

#calculating deployment issue for DA_Retention_ScPy vertical
deployment_retention_scpy_issue = calculating_issues(len(df), deployment_tag, retention_scpy_tag)
print("Deployment DA_Retention_ScPy Issue : ", deployment_retention_scpy_issue)

#calculating bydesign_valid for DA_Retention_ScPy vertical
bydesign_valid_retention_scpy_issue = calculating_issues(len(df), bydesign_valid_tag, retention_scpy_tag)
print("By Design Valid DA_Retention_ScPy Issue : ", bydesign_valid_retention_scpy_issue)

#calculating bydesign_tc for DA_Retention_ScPy vertical
bydesign_tc_retention_scpy_issue = calculating_issues(len(df), bydesign_tc_tag, retention_scpy_tag)
print("ByDesign TC DA_Retention_ScPy Issue : ", bydesign_tc_retention_scpy_issue)

#calculating business accepted defect for DA_Retention_ScPy vertical
business_accepted_defect_retention_scpy_issue = calculating_issues(len(df), business_accepted_defect_tag, retention_scpy_tag)
print("Business Accepted Defect for DA_Retention_ScPy : ", business_accepted_defect_retention_scpy_issue)

#calculating known issue for DA_Retention_ScPy vertical
knownissue_retention_scpy_issue = calculating_issues(len(df), knownissue_tag, retention_scpy_tag)
print("Known Issue for  DA_Retention_ScPy Vertical : ", knownissue_retention_scpy_issue)

#calulating blocker issue for DA_Retention_ScPy vertical
blocker_retention_scpy_issue = calculating_blocker_issues(len(df), retention_scpy_tag)
print("DA_Retention_ScPy Blocker Issues : ",blocker_retention_scpy_issue)

#calculating open and resolved blocker bugs for DA_Retention_ScPy vertical
blocker_retention_scpy_tag_open_issues, blocker_retention_scpy_tag_resolved_issues = calcualting_open_and_resolved_blocker_issues(len(df), retention_scpy_tag)
print("DA_Retention_ScPy Open Blocker Issues : ", blocker_retention_scpy_tag_open_issues)
print("DA_Retention_ScPy Resolved Blocker Issues : ", blocker_retention_scpy_tag_resolved_issues)

#calculating valid issues for DA_Retention_ScPy vertical
valid_retention_scpy_issue = (fixed_retention_scpy_issue+notfixed_valid_retention_scpy_issue+pdfixed_retention_scpy_issue+postlaunch_fix_retention_scpy_issue+ 
                              deferred_retention_scpy_issue+deployment_retention_scpy_issue+rcadispute_retention_scpy_issue+bydesign_valid_retention_scpy_issue+
                              bydesign_tc_retention_scpy_issue+business_accepted_defect_retention_scpy_issue)
print('Valid defects in DA_Retention_ScPy vertical : ',valid_retention_scpy_issue)

#valid defect ration for DA_Retention_ScPy Vertical
if resolved_retention_scpy_issue == 0:
    valid_defect_ratio_retention_scpy = 0
else:
    valid_defect_ratio_retention_scpy = int(valid_retention_scpy_issue*100/resolved_retention_scpy_issue)
print(str('Valid defect ration in DA_Retention_ScPy : ')+str(valid_defect_ratio_retention_scpy)+'%')

#defect fix ratio for DA_Retention_ScPy Vertical
if resolved_retention_scpy_issue == 0:
    defect_fix_ratio_retention_scpy = 0
else:
    defect_fix_ratio_retention_scpy = int(fixed_retention_scpy_issue*100/resolved_retention_scpy_issue)
print(str('Defect fix ration for DA_Retention_ScPy : ')+str(defect_fix_ratio_retention_scpy)+'%')

#writing the values to MBR Format excel sheet for DA_Retention_ScPy
wb = load_workbook(path_write_xl)
#selceting the active sheet from the excel file, active sheet is generally the sheet that we see when the excel file is opened
ws = wb.active
ws['D10'].value = retention_scpy_issues
ws['D11'].value = tc_retention_scpy_issue
ws['D12'].value = adhoc_retention_scpy_issue
ws['D13'].value = valid_retention_scpy_issue
ws['D14'].value = bydesign_tc_retention_scpy_issue+bydesign_valid_retention_scpy_issue
ws['D15'].value = rcadispute_acuistion_issue+deployment_retention_scpy_issue
ws['D16'].value = moved_to_tt_retention_scpy_issue
ws['D17'].value = business_accepted_defect_retention_scpy_issue+knownissue_retention_scpy_issue
ws['D18'].value = deployment_retention_scpy_issue
ws['D20'].value = cantreproduce_retention_scpy_issue
ws['D21'].value = duplicate_retention_scpy_issue
ws['D23'].value = fixed_retention_scpy_issue + pdfixed_retention_scpy_issue + postlaunch_fix_retention_scpy_issue
ws['D24'].value = resolved_retention_scpy_issue
ws['D25'].value = open_retention_scpy_issue
ws['D26'].value = blocker_retention_scpy_issue
ws['D27'].value = blocker_retention_scpy_tag_open_issues
ws['D28'].value = blocker_retention_scpy_tag_resolved_issues
ws['D34'].value = triaged_retention_scpy_issue
ws['D35'].value = valid_defect_ratio_retention_scpy
ws['D36'].value = defect_fix_ratio_retention_scpy
#closing the opened excel file after writing in it
wb.save(path_write_xl)

#calculating total DA_Retention_Speed QA issues
for i in range(0,len(df)):
    if retenetion_speedy_qa_tag in df.iloc[i,-1]:
        retenetion_speedy_qa_issues += 1
print('\n')
print("Total DA_Retention_Speed QA Issue : ",retenetion_speedy_qa_issues)

#calculating realoved issues for DA_Retention_Speed QA
for i in range(0,len(df)):
    if (('Closed' in df.iloc[i,3] or 'Resolved' in df.iloc[i,3]) and retenetion_speedy_qa_tag in df.iloc[i,-1]):
        resolved_retenetion_speedy_qa_issue += 1
print('Resolves issues for DA_Retention_Speed QA : ',resolved_retenetion_speedy_qa_issue)

#calculating open issues for DA_Retention_Speed QA issue
for i in range(1,len(df)):
    if 'Open' in df.iloc[i,3] and retenetion_speedy_qa_tag in df.iloc[i,-1]:
        open_retenetion_speedy_qa_issue += 1
print("Unresolved DA_Retention_Speed QA Issue : ",open_retenetion_speedy_qa_issue)

#calculating fixed issues for DA_Retention_Speed QA vertical
fixed_retenetion_speedy_qa_issue = calculating_issues(len(df), fixed_tag, retenetion_speedy_qa_tag)
print("Fixed DA_Retention_Speed QA Issues : ",fixed_retenetion_speedy_qa_issue)

#calculating cannot reproduce issues for DA_Retention_Speed QA Vertical
cantreproduce_retenetion_speedy_qa_issue = calculating_issues(len(df), cantreproduce_tag, retenetion_speedy_qa_tag)
print("Cannot Reproduce DA_Retention_Speed QA Issues : ",cantreproduce_retenetion_speedy_qa_issue)

#calculating duplicate issues for DA_Retention_Speed QA Vertical
duplicate_retenetion_speedy_qa_issue = calculating_issues(len(df), duplicate_tag, retenetion_speedy_qa_tag)
print("Duplicate DA_Retention_Speed QA Issues : ",duplicate_retenetion_speedy_qa_issue)

#calculating triaged issues for DA_Retention_Speed QA Vertical
triaged_retenetion_speedy_qa_issue = calculating_issues(len(df), triaged_tag, retenetion_speedy_qa_tag)
print("Triaged DA_Retention_Speed QA Issues : ",triaged_retenetion_speedy_qa_issue)

#calcualting deferred issues for DA_Retention_Speed QA Vertical
deferred_retenetion_speedy_qa_issue = calculating_issues(len(df), deferred_tag, retenetion_speedy_qa_tag)
print("Deferred DA_Retention_Speed QA Issue : ",deferred_retenetion_speedy_qa_issue)

#calculating rca_dispute issue for DA_Retention_Speed QA vertical
rcadispute_retenetion_speedy_qa_issue = calculating_issues(len(df), rcadispute_tag, retenetion_speedy_qa_tag)
print("RCA_Dispute DA_Retention_Speed QA Issue : ",rcadispute_retenetion_speedy_qa_issue)

#calculating tc issue for DA_Retention_Speed QA vertical
tc_retenetion_speedy_qa_issue = calculating_issues(len(df), tc_tag, retenetion_speedy_qa_tag)
print("TC DA_Retention_Speed QA Issue : ",tc_retenetion_speedy_qa_issue)

#calculating adhoc issue for DA_Retention_Speed QA vertical
adhoc_retenetion_speedy_qa_issue = calculating_issues(len(df), adhoc_tag, retenetion_speedy_qa_tag)
print("Adhoc DA_Retention_Speed QA Issue : ",adhoc_retenetion_speedy_qa_issue)

#calculating moved_to_tt_issues for DA_Retention_Speed QA vertical
moved_to_tt_retenetion_speedy_qa_issue = calculating_issues(len(df), moved_to_tt_tag, retenetion_speedy_qa_tag)
print("Moved to TT for DA_Retention_Speed QA: ", moved_to_tt_retenetion_speedy_qa_issue)

#calculating not_fixed_valid for DA_Retention_Speed QA vertical
notfixed_valid_retenetion_speedy_qa_issue = calculating_issues(len(df), notfixed_valid_tag, retenetion_speedy_qa_tag)
print("Not Fixed Valid DA_Retention_Speed QA Issue : ", notfixed_valid_retenetion_speedy_qa_issue)

#calculating pdfixed for DA_Retention_Speed QA vertical
pdfixed_retenetion_speedy_qa_issue = calculating_issues(len(df), pdfixed_tag, retenetion_speedy_qa_tag)
print("PD_Fixed DA_Retention_Speed QA Issue : ", pdfixed_retenetion_speedy_qa_issue)

#calculating postlaunch_fixed for DA_Retention_Speed QA vertical
postlaunch_fix_retenetion_speedy_qa_issue = calculating_issues(len(df), postlaunch_fix_tag, retenetion_speedy_qa_tag)
print("Post Launch Fix DA_Retention_Speed QA Issue : ", postlaunch_fix_retenetion_speedy_qa_issue)

#calculating deployment issue for DA_Retention_Speed QA vertical
deployment_retenetion_speedy_qa_issue = calculating_issues(len(df), deployment_tag, retenetion_speedy_qa_tag)
print("Deployment DA_Retention_Speed QA Issue : ", deployment_retenetion_speedy_qa_issue)

#calculating bydesign_valid for DA_Retention_Speed QA vertical
bydesign_valid_retenetion_speedy_qa_issue = calculating_issues(len(df), bydesign_valid_tag, retenetion_speedy_qa_tag)
print("By Design Valid DA_Retention_Speed QA Issue : ", bydesign_valid_retenetion_speedy_qa_issue)

#calculating bydesign_tc for DA_Retention_Speed QA vertical
bydesign_tc_retenetion_speedy_qa_issue = calculating_issues(len(df), bydesign_tc_tag, retenetion_speedy_qa_tag)
print("ByDesign TC DA_Retention_Speed QA Issue : ", bydesign_tc_retenetion_speedy_qa_issue)

#calculating business accepted defect for DA_Retention_Speed QA vertical
business_accepted_defect_retenetion_speedy_qa_issue = calculating_issues(len(df), business_accepted_defect_tag, retenetion_speedy_qa_tag)
print("Business Accepted Defect for DA_Retention_Speed QA : ", business_accepted_defect_retenetion_speedy_qa_issue)

#calculating known issue for DA_Retention_Speed QA vertical
knownissue_retenetion_speedy_qa_issue = calculating_issues(len(df), knownissue_tag, retenetion_speedy_qa_tag)
print("Known Issue for  DA_Retention_Speed QA Vertical : ", knownissue_retenetion_speedy_qa_issue)

#calulating blocker issue for DA_Retention_Speed QA vertical
blocker_issue_retenetion_speedy_qa = calculating_blocker_issues(len(df), retenetion_speedy_qa_tag)
print("DA_Retention_Speed QA Blocker Issues : ",blocker_issue_retenetion_speedy_qa)

#calculating open and resolved blocker bugs for DA_Retention_Speed QA vertical
blocker_open_issue_retenetion_speedy_qa, blocker_resolved_issue_retenetion_speedy_qa = calcualting_open_and_resolved_blocker_issues(len(df), retenetion_speedy_qa_tag)
print("DA_Retention_Speed QA Open Blocker Issues : ", blocker_open_issue_retenetion_speedy_qa)
print("DA_Retention_Speed QA Resolved Blocker Issues : ", blocker_resolved_issue_retenetion_speedy_qa)


#calculating valid issues for DA_Retention_Speed QA vertical
valid_retenetion_speedy_qa_issue = (fixed_retenetion_speedy_qa_issue+notfixed_valid_retenetion_speedy_qa_issue+pdfixed_retenetion_speedy_qa_issue+postlaunch_fix_retenetion_speedy_qa_issue+ 
                    deferred_retenetion_speedy_qa_issue+deployment_retenetion_speedy_qa_issue+rcadispute_retenetion_speedy_qa_issue+bydesign_valid_retenetion_speedy_qa_issue+
                    bydesign_tc_retenetion_speedy_qa_issue+business_accepted_defect_retenetion_speedy_qa_issue)
print('Valid defects in DA_Retention_Speed QA vertical : ',valid_retenetion_speedy_qa_issue)

#valid defect ration for DA_Retention_Speed QA Vertical
if resolved_retenetion_speedy_qa_issue == 0:
    valid_defect_ratio_retenetion_speedy_qa = 0
else:
    valid_defect_ratio_retenetion_speedy_qa = int(valid_retenetion_speedy_qa_issue*100/resolved_retenetion_speedy_qa_issue)
print(str('Valid defect ration in DA_Retention_Speed QA : ')+str(valid_defect_ratio_retenetion_speedy_qa)+'%')


#defect fix ratio for DA_Retention_Speed QA Vertical
if resolved_retenetion_speedy_qa_issue == 0:
    defect_fix_ratio_retenetion_speedy_qa = 0
else:
    defect_fix_ratio_retenetion_speedy_qa = int(fixed_retenetion_speedy_qa_issue*100/resolved_retenetion_speedy_qa_issue)
print(str('Defect fix ration for DA_Retention_Speed QA Central : ')+str(defect_fix_ratio_retenetion_speedy_qa)+'%')

#writing the values to MBR Format excel sheet for DA_Retention_Speed QA Vertical
wb = load_workbook(path_write_xl)
#selceting the active sheet from the excel file, active sheet is generally the sheet that we see when the excel file is opened
ws = wb.active
ws['E10'].value = retenetion_speedy_qa_issues
ws['E11'].value = tc_retenetion_speedy_qa_issue
ws['E12'].value = adhoc_retenetion_speedy_qa_issue
ws['E13'].value = valid_retenetion_speedy_qa_issue
ws['E14'].value = bydesign_tc_retenetion_speedy_qa_issue+bydesign_valid_retenetion_speedy_qa_issue
ws['E15'].value = rcadispute_acuistion_issue+deployment_retenetion_speedy_qa_issue
ws['E16'].value = moved_to_tt_retenetion_speedy_qa_issue
ws['E17'].value = business_accepted_defect_retenetion_speedy_qa_issue+knownissue_retenetion_speedy_qa_issue
ws['E18'].value = deployment_retenetion_speedy_qa_issue
ws['E20'].value = cantreproduce_retenetion_speedy_qa_issue
ws['E21'].value = duplicate_retenetion_speedy_qa_issue
ws['E23'].value = fixed_retenetion_speedy_qa_issue + pdfixed_retenetion_speedy_qa_issue + postlaunch_fix_retenetion_speedy_qa_issue
ws['E24'].value = resolved_retenetion_speedy_qa_issue
ws['E25'].value = open_retenetion_speedy_qa_issue
ws['E26'].value = blocker_issue_retenetion_speedy_qa
ws['E27'].value = blocker_open_issue_retenetion_speedy_qa
ws['E28'].value = blocker_resolved_issue_retenetion_speedy_qa
ws['E34'].value = triaged_retenetion_speedy_qa_issue
ws['E35'].value = valid_defect_ratio_retenetion_speedy_qa
ws['E36'].value = defect_fix_ratio_retenetion_speedy_qa
#closing the opened excel file after writing in it
wb.save(path_write_xl)

#calculating total DA_Payments Issues
for i in range(0,len(df)):
    if payments_tag in df.iloc[i,-1]:
        payments_issues += 1
print('\n')        
print("Total Pymants Issue : ",payments_issues)

#calculating total resolved DA_payments issues
for i in range(0,len(df)):
    if ('Closed' in df.iloc[i,3] or 'Resolved' in df.iloc[i,3]) and (payments_tag in df.iloc[i,-1]):
        resolved_payments_issue += 1
print('Resolved Payments Issue : ',resolved_payments_issue)

#calculating open issues for DA_payments issue
for i in range(1,len(df)):
    if 'Open' in df.iloc[i,3] and payments_tag in df.iloc[i,-1]:
        open_payments_issue += 1
print("Unresolved Payments Issue : ",open_payments_issue)

#calculating fixed issues for DA_payments vertical
fixed_payments_issue = calculating_issues(len(df), fixed_tag, payments_tag)
print("Fixed payments Issues : ",fixed_payments_issue)

#calculating cannot reproduce issues for DA_payments Vertical
cantreproduce_payments_issue = calculating_issues(len(df), cantreproduce_tag, payments_tag)
print("Cannot Reproduce payments Issues : ",cantreproduce_payments_issue)

#calculating duplicate issues for DA_payments Vertical
duplicate_payments_issue = calculating_issues(len(df), duplicate_tag, payments_tag)
print("Duplicate payments Issues : ",duplicate_payments_issue)

#calculating triaged issues for DA_payments Vertical
triaged_payments_issue = calculating_issues(len(df), triaged_tag, payments_tag)
print("Triaged payments Issues : ",triaged_payments_issue)

#calcualting deferred issues for DA_payments Vertical
deferred_payments_issue = calculating_issues(len(df), deferred_tag, payments_tag)
print("Deferred payments Issue : ",deferred_payments_issue)

#calculating rca_dispute issue for DA_payments vertical
rcadispute_payments_issue = calculating_issues(len(df), rcadispute_tag, payments_tag)
print("RCA_Dispute payments Issue : ",rcadispute_payments_issue)

#calculating tc issue for DA_payments vertical
tc_payments_issue = calculating_issues(len(df), tc_tag, payments_tag)
print("TC payments Issue : ",tc_payments_issue)

#calculating adhoc issue for DA_payments vertical
adhoc_payments_issue = calculating_issues(len(df), adhoc_tag, payments_tag)
print("Adhoc payments Issue : ",adhoc_payments_issue)

#calculating moved_to_tt_issues for DA_payments vertical
moved_to_tt_payments_issue = calculating_issues(len(df), moved_to_tt_tag, payments_tag)
print("Moved to TT for Payments : ", moved_to_tt_payments_issue)

#calculating not_fixed_valid for DA_payments vertical
notfixed_valid_payments_issue = calculating_issues(len(df), notfixed_valid_tag, payments_tag)
print("Not Fixed Valid Payments Issue : ", notfixed_valid_payments_issue)

#calculating pdfixed for DA_payments vertical
pdfixed_payments_issue = calculating_issues(len(df), pdfixed_tag, payments_tag)
print("PD_Fixed Payments Issue : ", pdfixed_payments_issue)

#calculating postlaunch_fixed for DA_payments vertical
postlaunch_fix_payments_issue = calculating_issues(len(df), postlaunch_fix_tag, payments_tag)
print("Post Launch Fix Payments Issue : ", postlaunch_fix_payments_issue)

#calculating deployment issue for DA_payments vertical
deployment_payments_issue = calculating_issues(len(df), deployment_tag, payments_tag)
print("Deployment Payments Issue : ", deployment_payments_issue)

#calculating bydesign_valid for DA_payments vertical
bydesign_valid_payments_issue = calculating_issues(len(df), bydesign_valid_tag, payments_tag)
print("By Design Valid Payments Issue : ", bydesign_valid_payments_issue)

#calculating bydesign_tc for DA_payments vertical
bydesign_tc_payments_issue = calculating_issues(len(df), bydesign_tc_tag, payments_tag)
print("ByDesign TC Payments Issue : ", bydesign_tc_payments_issue)

#calculating business accepted defect for DA_payments vertical
business_accepted_defect_payments_issue = calculating_issues(len(df), business_accepted_defect_tag, payments_tag)
print("Business Accepted Defect for Payments : ", business_accepted_defect_payments_issue)

#calculating known issue for DA_payments vertical
knownissue_payments_issue = calculating_issues(len(df), knownissue_tag, payments_tag)
print("Known Issue for  Payments Vertical : ", knownissue_payments_issue)

#calulating blocker issue for DA_payments vertical
blocker_payments_issue = calculating_blocker_issues(len(df), payments_tag)
print("DA_payments Blocker Issues : ",blocker_payments_issue)

#calculating open and resolved blocker bugs for DA_payments vertical
blocker_open_payments_issue, blocker_resolved_payments_issue = calcualting_open_and_resolved_blocker_issues(len(df), payments_tag)
print("DA_payments Open Blocker Issues : ", blocker_open_payments_issue)
print("DA_payments Resolved Blocker Issues : ", blocker_resolved_payments_issue)

#calculating valid issues for DA_payments vertical
valid_payments_issue = (fixed_payments_issue+notfixed_valid_payments_issue+pdfixed_payments_issue+postlaunch_fix_payments_issue+ 
                    deferred_payments_issue+deployment_payments_issue+rcadispute_payments_issue+bydesign_valid_payments_issue+
                    bydesign_tc_payments_issue+business_accepted_defect_payments_issue)
print('Valid defects in DA_payments vertical : ',valid_payments_issue)

#valid defect ration for DA_payments Vertical
if resolved_payments_issue == 0:
    valid_defect_ratio_payments = 0
else:
    valid_defect_ratio_payments = int(valid_payments_issue*100/resolved_payments_issue)
print(str('Valid defect ration in DA_payments : ')+str(valid_defect_ratio_payments)+'%')

#defect fix ratio for DA_payments Vertical
if resolved_payments_issue == 0:
    defect_fix_ratio_payments = 0
else:
    defect_fix_ratio_payments = int(fixed_payments_issue*100/resolved_payments_issue)
print(str('Defect fix ratio for payments : ')+str(defect_fix_ratio_payments)+'%')

#writing the values to MBR Format excel sheet for DA_Payments
wb = load_workbook(path_write_xl)
#selceting the active sheet from the excel file, active sheet is generally the sheet that we see when the excel file is opened
ws = wb.active
ws['F10'].value = payments_issues
ws['F11'].value = tc_payments_issue
ws['F12'].value = adhoc_payments_issue
ws['F13'].value = valid_payments_issue
ws['F14'].value = bydesign_tc_payments_issue+bydesign_valid_payments_issue
ws['F15'].value = rcadispute_acuistion_issue+deployment_payments_issue
ws['F16'].value = moved_to_tt_payments_issue
ws['F17'].value = business_accepted_defect_payments_issue+knownissue_payments_issue
ws['F18'].value = deployment_payments_issue
ws['F20'].value = cantreproduce_payments_issue
ws['F21'].value = duplicate_payments_issue
ws['F23'].value = fixed_payments_issue + pdfixed_payments_issue + postlaunch_fix_payments_issue
ws['F24'].value = resolved_payments_issue
ws['F25'].value = open_payments_issue
ws['F26'].value = blocker_payments_issue
ws['F27'].value = blocker_open_payments_issue
ws['F28'].value = blocker_resolved_payments_issue
ws['F34'].value = triaged_payments_issue
ws['F35'].value = valid_defect_ratio_payments
ws['F36'].value = defect_fix_ratio_payments
#closing the opened excel file after writing in it
wb.save(path_write_xl)

#calculating total DA_EU_MFA ISSUES
for i in range(0,len(df)):
    if eu_mfa_tag in df.iloc[i,-1]:
        eu_mfa_issues += 1
print("\n")
print("DA_EU_MFA Issues : ",eu_mfa_issues)

#calculating resolved issues for DA_EU_MFA
for i in range(0,len(df)):
    if(('Closed' in df.iloc[i,3] or 'Resolved' in df.iloc[i,3]) and eu_mfa_tag in df.iloc[i,-1]):
        resolved_eu_mfa_issue += 1
print('DA_EU_MFA Resolved Issues : ',resolved_eu_mfa_issue)

#calculating open issues for DA_EU_MFA issue
for i in range(1,len(df)):
    if 'Open' in df.iloc[i,3] and eu_mfa_tag in df.iloc[i,-1]:
        open_eu_mfa_issue += 1
print("Unresolved DA_EU_MFA Issue : ",open_eu_mfa_issue)

#calculating fixed issues for DA_EU_MFA vertical
fixed_eu_mfa_issue = calculating_issues(len(df), fixed_tag, eu_mfa_tag)
print("Fixed EU MFA Issues : ",fixed_eu_mfa_issue)

#calculating cannot reproduce issues for DA_EU_MFA Vertical
cantreproduce_eu_mfa_issue = calculating_issues(len(df), cantreproduce_tag, eu_mfa_tag)
print("Cannot Reproduce EU MFA Issues : ",cantreproduce_eu_mfa_issue)

#calculating duplicate issues for DA_EU_MFA Vertical
duplicate_eu_mfa_issue = calculating_issues(len(df), duplicate_tag, eu_mfa_tag)
print("Duplicate EU MFA Issues : ",duplicate_eu_mfa_issue)

#calculating triaged issues for DA_EU_MFA Vertical
triaged_eu_mfa_issue = calculating_issues(len(df), triaged_tag, eu_mfa_tag)
print("Triaged EU MFA Issues : ",triaged_eu_mfa_issue)

#calcualting deferred issues for DA_EU_MFA Vertical
deferred_eu_mfa_issue = calculating_issues(len(df), deferred_tag, eu_mfa_tag)
print("Deferred EU MFA Issue : ",deferred_eu_mfa_issue)

#calculating rca_dispute issue for DA_EU_MFA vertical
rcadispute_eu_mfa_issue = calculating_issues(len(df), rcadispute_tag, eu_mfa_tag)
print("RCA_Dispute EU MFA Issue : ",rcadispute_eu_mfa_issue)

#calculating tc issue for DA_EU_MFA vertical
tc_eu_mfa_issue = calculating_issues(len(df), tc_tag, eu_mfa_tag)
print("TC EU MFA Issue : ",tc_eu_mfa_issue)

#calculating adhoc issue for DA_EU_MFA vertical
adhoc_eu_mfa_issue = calculating_issues(len(df), adhoc_tag, eu_mfa_tag)
print("Adhoc EU MFA Issue : ",adhoc_eu_mfa_issue)

#calculating moved_to_tt_issues for DA_EU_MFA vertical
moved_to_tt_eu_mfa_issue = calculating_issues(len(df), moved_to_tt_tag, eu_mfa_tag)
print("Moved to TT for EU MFA: ", moved_to_tt_eu_mfa_issue)

#calculating not_fixed_valid for DA_EU_MFA vertical
notfixed_valid_eu_mfa_issue = calculating_issues(len(df), notfixed_valid_tag, eu_mfa_tag)
print("Not Fixed Valid EU MFA Issue : ", notfixed_valid_eu_mfa_issue)

#calculating pdfixed for DA_EU_MFA vertical
pdfixed_eu_mfa_issue = calculating_issues(len(df), pdfixed_tag, eu_mfa_tag)
print("PD_Fixed EU MFA Issue : ", pdfixed_eu_mfa_issue)

#calculating postlaunch_fixed for DA_EU_MFA vertical
postlaunch_fix_eu_mfa_issue = calculating_issues(len(df), postlaunch_fix_tag, eu_mfa_tag)
print("Post Launch Fix EU MFA Issue : ", postlaunch_fix_eu_mfa_issue)

#calculating deployment issue for DA_EU_MFA vertical
deployment_eu_mfa_issue = calculating_issues(len(df), deployment_tag, eu_mfa_tag)
print("Deployment EU MFA Issue : ", deployment_eu_mfa_issue)

#calculating bydesign_valid for DA_EU_MFA vertical
bydesign_valid_eu_mfa_issue = calculating_issues(len(df), bydesign_valid_tag, eu_mfa_tag)
print("By Design Valid EU MFA Issue : ", bydesign_valid_eu_mfa_issue)

#calculating bydesign_tc for DA_EU_MFA vertical
bydesign_tc_eu_mfa_issue = calculating_issues(len(df), bydesign_tc_tag, eu_mfa_tag)
print("ByDesign TC EU MFA Issue : ", bydesign_tc_eu_mfa_issue)

#calculating business accepted defect for DA_EU_MFA vertical
business_accepted_defect_eu_mfa_issue = calculating_issues(len(df), business_accepted_defect_tag, eu_mfa_tag)
print("Business Accepted Defect for EU MFA : ", business_accepted_defect_eu_mfa_issue)

#calculating known issue for DA_EU_MFA vertical
knownissue_eu_mfa_issue = calculating_issues(len(df), knownissue_tag, eu_mfa_tag)
print("Known Issue for  EU MFA Vertical : ", knownissue_eu_mfa_issue)

#calulating blocker issue for DA_EU_MFA vertical
blocker_eu_mfa_issue = calculating_blocker_issues(len(df), eu_mfa_tag)
print("DA_EU_MFA Blocker Issues : ",blocker_eu_mfa_issue)

#calculating open and resolved blocker bugs for DA_EU_MFA vertical
blocker_open_eu_mfa_issue, blocker_resolved_eu_mfa_issue = calcualting_open_and_resolved_blocker_issues(len(df), eu_mfa_tag)
print("DA_EU_MFA Open Blocker Issues : ", blocker_open_eu_mfa_issue)
print("DA_EU_MFA Resolved Blocker Issues : ", blocker_resolved_eu_mfa_issue)

#calculating valid issues for DA_EU_MFA vertical
valid_eu_mfa_issue = (fixed_eu_mfa_issue+notfixed_valid_eu_mfa_issue+pdfixed_eu_mfa_issue+postlaunch_fix_eu_mfa_issue+ 
                    deferred_eu_mfa_issue+deployment_eu_mfa_issue+rcadispute_eu_mfa_issue+bydesign_valid_eu_mfa_issue+
                    bydesign_tc_eu_mfa_issue+business_accepted_defect_eu_mfa_issue)
print('Valid defects in DA_EU_MFA vertical : ',valid_eu_mfa_issue)

#valid defect ration for DA_EU_MFA Vertical
if resolved_eu_mfa_issue == 0:
    valid_defect_ratio_eu_mfa = 0
else:
    valid_defect_ratio_eu_mfa = int(valid_eu_mfa_issue*100/resolved_eu_mfa_issue)
print(str('Valid defect ration in DA_EU_MFA : ')+str(valid_defect_ratio_eu_mfa)+'%')


#defect fix ratio for DA_EU_MFA Vertical
if resolved_eu_mfa_issue == 0:
    defect_fix_ratio_eu_mfa = 0
else:
    defect_fix_ratio_eu_mfa = int(fixed_eu_mfa_issue*100/resolved_eu_mfa_issue)
print(str('Defect fix ration for EU MFA : ')+str(defect_fix_ratio_eu_mfa)+'%')

#writing the values to MBR Format excel sheet for DA_EU_MFA
wb = load_workbook(path_write_xl)
#selceting the active sheet from the excel file, active sheet is generally the sheet that we see when the excel file is opened
ws = wb.active
ws['G10'].value = eu_mfa_issues
ws['G11'].value = tc_eu_mfa_issue
ws['G12'].value = adhoc_eu_mfa_issue
ws['G13'].value = valid_eu_mfa_issue
ws['G14'].value = bydesign_tc_eu_mfa_issue+bydesign_valid_eu_mfa_issue
ws['G15'].value = rcadispute_acuistion_issue+deployment_eu_mfa_issue
ws['G16'].value = moved_to_tt_eu_mfa_issue
ws['G17'].value = business_accepted_defect_eu_mfa_issue+knownissue_eu_mfa_issue
ws['G18'].value = deployment_eu_mfa_issue
ws['G20'].value = cantreproduce_eu_mfa_issue
ws['G21'].value = duplicate_eu_mfa_issue
ws['G23'].value = fixed_eu_mfa_issue + pdfixed_eu_mfa_issue + postlaunch_fix_eu_mfa_issue
ws['G24'].value = resolved_eu_mfa_issue
ws['G25'].value = open_eu_mfa_issue
ws['G26'].value = blocker_eu_mfa_issue
ws['G27'].value = blocker_open_eu_mfa_issue
ws['G28'].value = blocker_resolved_eu_mfa_issue
ws['G34'].value = triaged_eu_mfa_issue
ws['G35'].value = valid_defect_ratio_eu_mfa
ws['G36'].value = defect_fix_ratio_eu_mfa
#closing the opened excel file after writing in it
wb.save(path_write_xl)

#calculating total DA_HVE issues
for i in range(0,len(df)):
    if hve_tag in df.iloc[i,-1]:
        hve_issues += 1
print('\n')
print("Total DA_HVE Issue : ",hve_issues)

#calculating realoved issues for DA_HVE
for i in range(0,len(df)):
    if (('Closed' in df.iloc[i,3] or 'Resolved' in df.iloc[i,3]) and hve_tag in df.iloc[i,-1]):
        resolved_hve_issue += 1
print('Resolves issues for DA_HVE : ',resolved_hve_issue)

#calculating open issues for DA_HVE issue
for i in range(1,len(df)):
    if 'Open' in df.iloc[i,3] and hve_tag in df.iloc[i,-1]:
        open_hve_issue += 1
print("Unresolved DA_HVE Issue : ",open_hve_issue)

#calculating fixed issues for DA_HVE vertical
fixed_hve_issue = calculating_issues(len(df), fixed_tag, hve_tag)
print("Fixed DA_HVE Issues : ",fixed_hve_issue)

#calculating cannot reproduce issues for DA_HVE Vertical
cantreproduce_hve_issue = calculating_issues(len(df), cantreproduce_tag, hve_tag)
print("Cannot Reproduce DA_HVE Issues : ",cantreproduce_hve_issue)

#calculating duplicate issues for DA_HVE Vertical
duplicate_hve_issue = calculating_issues(len(df), duplicate_tag, hve_tag)
print("Duplicate DA_HVE Issues : ",duplicate_hve_issue)

#calculating triaged issues for DA_HVE Vertical
triaged_hve_issue = calculating_issues(len(df), triaged_tag, hve_tag)
print("Triaged DA_HVE Issues : ",triaged_hve_issue)

#calcualting deferred issues for DA_HVE Vertical
deferred_hve_issue = calculating_issues(len(df), deferred_tag, hve_tag)
print("Deferred DA_HVE Issue : ",deferred_hve_issue)

#calculating rca_dispute issue for DA_HVE vertical
rcadispute_hve_issue = calculating_issues(len(df), rcadispute_tag, hve_tag)
print("RCA_Dispute DA_HVE Issue : ",rcadispute_hve_issue)

#calculating tc issue for DA_HVE vertical
tc_hve_issue = calculating_issues(len(df), tc_tag, hve_tag)
print("TC DA_HVE Issue : ",tc_hve_issue)

#calculating adhoc issue for DA_HVE vertical
adhoc_hve_issue = calculating_issues(len(df), adhoc_tag, hve_tag)
print("Adhoc DA_HVE Issue : ",adhoc_hve_issue)

#calculating moved_to_tt_issues for DA_HVE vertical
moved_to_tt_hve_issue = calculating_issues(len(df), moved_to_tt_tag, hve_tag)
print("Moved to TT for DA_HVE: ", moved_to_tt_hve_issue)

#calculating not_fixed_valid for DA_HVE vertical
notfixed_valid_hve_issue = calculating_issues(len(df), notfixed_valid_tag, hve_tag)
print("Not Fixed Valid DA_HVE Issue : ", notfixed_valid_hve_issue)

#calculating pdfixed for DA_HVE vertical
pdfixed_hve_issue = calculating_issues(len(df), pdfixed_tag, hve_tag)
print("PD_Fixed DA_HVE Issue : ", pdfixed_hve_issue)

#calculating postlaunch_fixed for DA_HVE vertical
postlaunch_fix_hve_issue = calculating_issues(len(df), postlaunch_fix_tag, hve_tag)
print("Post Launch Fix DA_HVE Issue : ", postlaunch_fix_hve_issue)

#calculating deployment issue for DA_HVE vertical
deployment_hve_issue = calculating_issues(len(df), deployment_tag, hve_tag)
print("Deployment DA_HVE Issue : ", deployment_hve_issue)

#calculating bydesign_valid for DA_HVE vertical
bydesign_valid_hve_issue = calculating_issues(len(df), bydesign_valid_tag, hve_tag)
print("By Design Valid DA_HVE Issue : ", bydesign_valid_hve_issue)

#calculating bydesign_tc for DA_HVE vertical
bydesign_tc_hve_issue = calculating_issues(len(df), bydesign_tc_tag, hve_tag)
print("ByDesign TC DA_HVE Issue : ", bydesign_tc_hve_issue)

#calculating business accepted defect for DA_HVE vertical
business_accepted_defect_hve_issue = calculating_issues(len(df), business_accepted_defect_tag, hve_tag)
print("Business Accepted Defect for DA_HVE : ", business_accepted_defect_hve_issue)

#calculating known issue for DA_HVE vertical
knownissue_hve_issue = calculating_issues(len(df), knownissue_tag, hve_tag)
print("Known Issue for  DA_HVE Vertical : ", knownissue_hve_issue)

#calulating blocker issue for DA_HVE vertical
blocker_hve_issues = calculating_blocker_issues(len(df), hve_tag)
print("DA_HVE Blocker Issues : ",blocker_hve_issues)

#calculating open and resolved blocker bugs for DA_HVE vertical
blocker_open_hve_issues, blocker_resolved_hve_issues = calcualting_open_and_resolved_blocker_issues(len(df), hve_tag)
print("DA_HVE Open Blocker Issues : ", blocker_open_hve_issues)
print("DA_HVE Resolved Blocker Issues : ", blocker_resolved_hve_issues)

#calculating valid issues for DA_HVE vertical
valid_hve_issue = (fixed_hve_issue+notfixed_valid_hve_issue+pdfixed_hve_issue+postlaunch_fix_hve_issue+ 
                    deferred_hve_issue+deployment_hve_issue+rcadispute_hve_issue+bydesign_valid_hve_issue+
                    bydesign_tc_hve_issue+business_accepted_defect_hve_issue)
print('Valid defects in DA_HVE vertical : ',valid_hve_issue)

#valid defect ration for DA_HVE Vertical
if resolved_hve_issue == 0:
    valid_defect_ratio_hve = 0
else:
    valid_defect_ratio_hve = int(valid_hve_issue*100/resolved_hve_issue)
print(str('Valid defect ration in DA_HVE : ')+str(valid_defect_ratio_hve)+'%')


#defect fix ratio for DA_HVE Vertical
if resolved_hve_issue == 0:
    defect_fix_ratio_hve = 0
else:
    defect_fix_ratio_hve = int(fixed_hve_issue*100/resolved_hve_issue)
print(str('Defect fix ration for DA_HVE Central : ')+str(defect_fix_ratio_hve)+'%')

#writing the values to MBR Format excel sheet for DA_HVE Vertical
wb = load_workbook(path_write_xl)
#selceting the active sheet from the excel file, active sheet is generally the sheet that we see when the excel file is opened
ws = wb.active
ws['H10'].value = hve_issues
ws['H11'].value = tc_hve_issue
ws['H12'].value = adhoc_hve_issue
ws['H13'].value = valid_hve_issue
ws['H14'].value = bydesign_tc_hve_issue+bydesign_valid_hve_issue
ws['H15'].value = rcadispute_acuistion_issue+deployment_hve_issue
ws['H16'].value = moved_to_tt_hve_issue
ws['H17'].value = business_accepted_defect_hve_issue+knownissue_hve_issue
ws['H18'].value = deployment_hve_issue
ws['H20'].value = cantreproduce_hve_issue
ws['H21'].value = duplicate_hve_issue
ws['H23'].value = fixed_hve_issue + pdfixed_hve_issue + postlaunch_fix_hve_issue
ws['H24'].value = resolved_hve_issue
ws['H25'].value = open_hve_issue
ws['H26'].value = blocker_hve_issues
ws['H27'].value = blocker_open_hve_issues
ws['H28'].value = blocker_resolved_hve_issues
ws['H34'].value = triaged_hve_issue
ws['H35'].value = valid_defect_ratio_hve
ws['H36'].value = defect_fix_ratio_hve
#closing the opened excel file after writing in it
wb.save(path_write_xl)

#calculating total DA_Shopping issues
for i in range(0,len(df)):
    if shopping_tag in df.iloc[i,-1]:
        shopping_issues += 1
print('\n')
print("Total DA_Shopping Issue : ",shopping_issues)

#calculating realoved issues for DA_Shopping
for i in range(0,len(df)):
    if (('Closed' in df.iloc[i,3] or 'Resolved' in df.iloc[i,3]) and shopping_tag in df.iloc[i,-1]):
        resolved_shopping_issue += 1
print('Resolves issues for DA_Shopping : ',resolved_shopping_issue)

#calculating open issues for DA_Shopping issue
for i in range(1,len(df)):
    if 'Open' in df.iloc[i,3] and shopping_tag in df.iloc[i,-1]:
        open_shopping_issue += 1
print("Unresolved DA_Shopping Issue : ",open_shopping_issue)

#calculating fixed issues for DA_Shopping vertical
fixed_shopping_issue = calculating_issues(len(df), fixed_tag, shopping_tag)
print("Fixed DA_Shopping Issues : ",fixed_shopping_issue)

#calculating cannot reproduce issues for DA_Shopping Vertical
cantreproduce_shopping_issue = calculating_issues(len(df), cantreproduce_tag, shopping_tag)
print("Cannot Reproduce DA_Shopping Issues : ",cantreproduce_shopping_issue)

#calculating duplicate issues for DA_Shopping Vertical
duplicate_shopping_issue = calculating_issues(len(df), duplicate_tag, shopping_tag)
print("Duplicate DA_Shopping Issues : ",duplicate_shopping_issue)

#calculating triaged issues for DA_Shopping Vertical
triaged_shopping_issue = calculating_issues(len(df), triaged_tag, shopping_tag)
print("Triaged DA_Shopping Issues : ",triaged_shopping_issue)

#calcualting deferred issues for DA_Shopping Vertical
deferred_shopping_issue = calculating_issues(len(df), deferred_tag, shopping_tag)
print("Deferred DA_Shopping Issue : ",deferred_shopping_issue)

#calculating rca_dispute issue for DA_Shopping vertical
rcadispute_shopping_issue = calculating_issues(len(df), rcadispute_tag, shopping_tag)
print("RCA_Dispute DA_Shopping Issue : ",rcadispute_shopping_issue)

#calculating tc issue for DA_Shopping vertical
tc_shopping_issue = calculating_issues(len(df), tc_tag, shopping_tag)
print("TC DA_Shopping Issue : ",tc_shopping_issue)

#calculating adhoc issue for DA_Shopping vertical
adhoc_shopping_issue = calculating_issues(len(df), adhoc_tag, shopping_tag)
print("Adhoc DA_Shopping Issue : ",adhoc_shopping_issue)

#calculating moved_to_tt_issues for DA_Shopping vertical
moved_to_tt_shopping_issue = calculating_issues(len(df), moved_to_tt_tag, shopping_tag)
print("Moved to TT for DA_Shopping: ", moved_to_tt_shopping_issue)

#calculating not_fixed_valid for DA_Shopping vertical
notfixed_valid_shopping_issue = calculating_issues(len(df), notfixed_valid_tag, shopping_tag)
print("Not Fixed Valid DA_Shopping Issue : ", notfixed_valid_shopping_issue)

#calculating pdfixed for DA_Shopping vertical
pdfixed_shopping_issue = calculating_issues(len(df), pdfixed_tag, shopping_tag)
print("PD_Fixed DA_Shopping Issue : ", pdfixed_shopping_issue)

#calculating postlaunch_fixed for DA_Shopping vertical
postlaunch_fix_shopping_issue = calculating_issues(len(df), postlaunch_fix_tag, shopping_tag)
print("Post Launch Fix DA_Shopping Issue : ", postlaunch_fix_shopping_issue)

#calculating deployment issue for DA_Shopping vertical
deployment_shopping_issue = calculating_issues(len(df), deployment_tag, shopping_tag)
print("Deployment DA_Shopping Issue : ", deployment_shopping_issue)

#calculating bydesign_valid for DA_Shopping vertical
bydesign_valid_shopping_issue = calculating_issues(len(df), bydesign_valid_tag, shopping_tag)
print("By Design Valid DA_Shopping Issue : ", bydesign_valid_shopping_issue)

#calculating bydesign_tc for DA_Shopping vertical
bydesign_tc_shopping_issue = calculating_issues(len(df), bydesign_tc_tag, shopping_tag)
print("ByDesign TC DA_Shopping Issue : ", bydesign_tc_shopping_issue)

#calculating business accepted defect for DA_Shopping vertical
business_accepted_defect_shopping_issue = calculating_issues(len(df), business_accepted_defect_tag, shopping_tag)
print("Business Accepted Defect for DA_Shopping : ", business_accepted_defect_shopping_issue)

#calculating known issue for DA_Shopping vertical
knownissue_shopping_issue = calculating_issues(len(df), knownissue_tag, shopping_tag)
print("Known Issue for  DA_Shopping Vertical : ", knownissue_shopping_issue)

#calulating blocker issue for DA_Shopping vertical
blocker_shopping_issue = calculating_blocker_issues(len(df), shopping_tag)
print("DA_Shopping Blocker Issues : ",blocker_shopping_issue)

#calculating open and resolved blocker bugs for DA_Shopping vertical
blocker_open_shopping_issue, blocker_resolved_shopping_issue = calcualting_open_and_resolved_blocker_issues(len(df), shopping_tag)
print("DA_Shopping Open Blocker Issues : ", blocker_open_shopping_issue)
print("DA_Shopping Resolved Blocker Issues : ", blocker_resolved_shopping_issue)

#calculating valid issues for DA_Shopping vertical
valid_shopping_issue = (fixed_shopping_issue+notfixed_valid_shopping_issue+pdfixed_shopping_issue+postlaunch_fix_shopping_issue+ 
                        deferred_shopping_issue+deployment_shopping_issue+rcadispute_shopping_issue+bydesign_valid_shopping_issue+
                        bydesign_tc_shopping_issue+business_accepted_defect_shopping_issue)
print('Valid defects in DA_Shopping vertical : ',valid_shopping_issue)

#valid defect ration for DA_Shopping Vertical
if resolved_shopping_issue == 0:
    valid_defect_ratio_shopping = 0
else:
    valid_defect_ratio_shopping = int(valid_shopping_issue*100/resolved_shopping_issue)
print(str('Valid defect ration in DA_Shopping : ')+str(valid_defect_ratio_shopping)+'%')


#defect fix ratio for DA_Shopping Vertical
if resolved_shopping_issue == 0:
    defect_fix_ratio_shopping = 0
else:
    defect_fix_ratio_shopping = int(fixed_shopping_issue*100/resolved_shopping_issue)
print(str('Defect fix ration for DA_Shopping : ')+str(defect_fix_ratio_shopping)+'%')

#writing the values to MBR Format excel sheet for DA_Shopping
wb = load_workbook(path_write_xl)
#selceting the active sheet from the excel file, active sheet is generally the sheet that we see when the excel file is opened
ws = wb.active
ws['I10'].value = shopping_issues
ws['I11'].value = tc_shopping_issue
ws['I12'].value = adhoc_shopping_issue
ws['I13'].value = valid_shopping_issue
ws['I14'].value = bydesign_tc_shopping_issue+bydesign_valid_shopping_issue
ws['I15'].value = rcadispute_acuistion_issue+deployment_shopping_issue
ws['I16'].value = moved_to_tt_shopping_issue
ws['I17'].value = business_accepted_defect_shopping_issue+knownissue_shopping_issue
ws['I18'].value = deployment_shopping_issue
ws['I20'].value = cantreproduce_shopping_issue
ws['I21'].value = duplicate_shopping_issue
ws['I23'].value = fixed_shopping_issue + pdfixed_shopping_issue + postlaunch_fix_shopping_issue
ws['I24'].value = resolved_shopping_issue
ws['I25'].value = open_shopping_issue
ws['I26'].value = blocker_shopping_issue
ws['I27'].value = blocker_open_shopping_issue
ws['I28'].value = blocker_resolved_shopping_issue
ws['I34'].value = triaged_shopping_issue
ws['I35'].value = valid_defect_ratio_shopping
ws['I36'].value = defect_fix_ratio_shopping
#closing the opened excel file after writing in it
wb.save(path_write_xl)

#calculating total DA_Retention_FDR issues
for i in range(0,len(df)):
    if retenetion_fdr_tag in df.iloc[i,-1]:
        retenetion_fdr_issues += 1
print('\n')
print("Total DA_Retention_FDR Issue : ",retenetion_fdr_issues)

#calculating realoved issues for DA_Retention_FDR
for i in range(0,len(df)):
    if (('Closed' in df.iloc[i,3] or 'Resolved' in df.iloc[i,3]) and retenetion_fdr_tag in df.iloc[i,-1]):
        resolved_retenetion_fdr_issue += 1
print('Resolves issues for DA_Retention_FDR : ',resolved_retenetion_fdr_issue)

#calculating open issues for DA_Retention_FDR issue
for i in range(1,len(df)):
    if 'Open' in df.iloc[i,3] and retenetion_fdr_tag in df.iloc[i,-1]:
        open_retenetion_fdr_issue += 1
print("Unresolved DA_Retention_FDR Issue : ",open_retenetion_fdr_issue)

#calculating fixed issues for DA_Retention_FDR vertical
fixed_retenetion_fdr_issue = calculating_issues(len(df), fixed_tag, retenetion_fdr_tag)
print("Fixed DA_Retention_FDR Issues : ",fixed_retenetion_fdr_issue)

#calculating cannot reproduce issues for DA_Retention_FDR Vertical
cantreproduce_retenetion_fdr_issue = calculating_issues(len(df), cantreproduce_tag, retenetion_fdr_tag)
print("Cannot Reproduce DA_Retention_FDR Issues : ",cantreproduce_retenetion_fdr_issue)

#calculating duplicate issues for DA_Retention_FDR Vertical
duplicate_retenetion_fdr_issue = calculating_issues(len(df), duplicate_tag, retenetion_fdr_tag)
print("Duplicate DA_Retention_FDR Issues : ",duplicate_retenetion_fdr_issue)

#calculating triaged issues for DA_Retention_FDR Vertical
triaged_retenetion_fdr_issue = calculating_issues(len(df), triaged_tag, retenetion_fdr_tag)
print("Triaged DA_Retention_FDR Issues : ",triaged_retenetion_fdr_issue)

#calcualting deferred issues for DA_Retention_FDR Vertical
deferred_retenetion_fdr_issue = calculating_issues(len(df), deferred_tag, retenetion_fdr_tag)
print("Deferred DA_Retention_FDR Issue : ",deferred_retenetion_fdr_issue)

#calculating rca_dispute issue for DA_Retention_FDR vertical
rcadispute_retenetion_fdr_issue = calculating_issues(len(df), rcadispute_tag, retenetion_fdr_tag)
print("RCA_Dispute DA_Retention_FDR Issue : ",rcadispute_retenetion_fdr_issue)

#calculating tc issue for DA_Retention_FDR vertical
tc_retenetion_fdr_issue = calculating_issues(len(df), tc_tag, retenetion_fdr_tag)
print("TC DA_Retention_FDR Issue : ",tc_retenetion_fdr_issue)

#calculating adhoc issue for DA_Retention_FDR vertical
adhoc_retenetion_fdr_issue = calculating_issues(len(df), adhoc_tag, retenetion_fdr_tag)
print("Adhoc DA_Retention_FDR Issue : ",adhoc_retenetion_fdr_issue)

#calculating moved_to_tt_issues for DA_Retention_FDR vertical
moved_to_tt_retenetion_fdr_issue = calculating_issues(len(df), moved_to_tt_tag, retenetion_fdr_tag)
print("Moved to TT for DA_Retention_FDR: ", moved_to_tt_retenetion_fdr_issue)

#calculating not_fixed_valid for DA_Retention_FDR vertical
notfixed_valid_retenetion_fdr_issue = calculating_issues(len(df), notfixed_valid_tag, retenetion_fdr_tag)
print("Not Fixed Valid DA_Retention_FDR Issue : ", notfixed_valid_retenetion_fdr_issue)

#calculating pdfixed for DA_Retention_FDR vertical
pdfixed_retenetion_fdr_issue = calculating_issues(len(df), pdfixed_tag, retenetion_fdr_tag)
print("PD_Fixed DA_Retention_FDR Issue : ", pdfixed_retenetion_fdr_issue)

#calculating postlaunch_fixed for DA_Retention_FDR vertical
postlaunch_fix_retenetion_fdr_issue = calculating_issues(len(df), postlaunch_fix_tag, retenetion_fdr_tag)
print("Post Launch Fix DA_Retention_FDR Issue : ", postlaunch_fix_retenetion_fdr_issue)

#calculating deployment issue for DA_Retention_FDR vertical
deployment_retenetion_fdr_issue = calculating_issues(len(df), deployment_tag, retenetion_fdr_tag)
print("Deployment DA_Retention_FDR Issue : ", deployment_retenetion_fdr_issue)

#calculating bydesign_valid for DA_Retention_FDR vertical
bydesign_valid_retenetion_fdr_issue = calculating_issues(len(df), bydesign_valid_tag, retenetion_fdr_tag)
print("By Design Valid DA_Retention_FDR Issue : ", bydesign_valid_retenetion_fdr_issue)

#calculating bydesign_tc for DA_Retention_FDR vertical
bydesign_tc_retenetion_fdr_issue = calculating_issues(len(df), bydesign_tc_tag, retenetion_fdr_tag)
print("ByDesign TC DA_Retention_FDR Issue : ", bydesign_tc_retenetion_fdr_issue)

#calculating business accepted defect for DA_Retention_FDR vertical
business_accepted_defect_retenetion_fdr_issue = calculating_issues(len(df), business_accepted_defect_tag, retenetion_fdr_tag)
print("Business Accepted Defect for DA_Retention_FDR : ", business_accepted_defect_retenetion_fdr_issue)

#calculating known issue for DA_Retention_FDR vertical
knownissue_retenetion_fdr_issue = calculating_issues(len(df), knownissue_tag, retenetion_fdr_tag)
print("Known Issue for  DA_Retention_FDR Vertical : ", knownissue_retenetion_fdr_issue)

#calulating blocker issue for DA_Retention_FDR vertical
blocker_retenetion_fdr_issue = calculating_blocker_issues(len(df), retenetion_fdr_tag)
print("DA_Retention_FDR Blocker Issues : ",blocker_retenetion_fdr_issue)

#calculating open and resolved blocker bugs for DA_Retention_FDR vertical
blocker_open_retenetion_fdr_issue, blocker_resolved_retenetion_fdr_issue = calcualting_open_and_resolved_blocker_issues(len(df), retenetion_fdr_tag)
print("DA_Retention_FDR Open Blocker Issues : ", blocker_open_retenetion_fdr_issue)
print("DA_Retention_FDR Resolved Blocker Issues : ", blocker_resolved_retenetion_fdr_issue)

#calculating valid issues for DA_Retention_FDR vertical
valid_retenetion_fdr_issue = (fixed_retenetion_fdr_issue+notfixed_valid_retenetion_fdr_issue+pdfixed_retenetion_fdr_issue+postlaunch_fix_retenetion_fdr_issue+ 
                    deferred_retenetion_fdr_issue+deployment_retenetion_fdr_issue+rcadispute_retenetion_fdr_issue+bydesign_valid_retenetion_fdr_issue+
                    bydesign_tc_retenetion_fdr_issue+business_accepted_defect_retenetion_fdr_issue)
print('Valid defects in DA_Retention_FDR vertical : ',valid_retenetion_fdr_issue)

#valid defect ration for DA_Retention_FDR Vertical
if resolved_retenetion_fdr_issue == 0:
    valid_defect_ratio_retenetion_fdr = 0
else:
    valid_defect_ratio_retenetion_fdr = int(valid_retenetion_fdr_issue*100/resolved_retenetion_fdr_issue)
print(str('Valid defect ration in DA_Retention_FDR : ')+str(valid_defect_ratio_retenetion_fdr)+'%')


#defect fix ratio for DA_Retention_FDR Vertical
if resolved_retenetion_fdr_issue == 0:
    defect_fix_ratio_retenetion_fdr = 0
else:
    defect_fix_ratio_retenetion_fdr = int(fixed_retenetion_fdr_issue*100/resolved_retenetion_fdr_issue)
print(str('Defect fix ration for DA_Retention_FDR Central : ')+str(defect_fix_ratio_retenetion_fdr)+'%')

#writing the values to MBR Format excel sheet for DA_Retention_FDR Vertical
wb = load_workbook(path_write_xl)
#selceting the active sheet from the excel file, active sheet is generally the sheet that we see when the excel file is opened
ws = wb.active
ws['J10'].value = retenetion_fdr_issues
ws['J11'].value = tc_retenetion_fdr_issue
ws['J12'].value = adhoc_retenetion_fdr_issue
ws['J13'].value = valid_retenetion_fdr_issue
ws['J14'].value = bydesign_tc_retenetion_fdr_issue+bydesign_valid_retenetion_fdr_issue
ws['J15'].value = rcadispute_acuistion_issue+deployment_retenetion_fdr_issue
ws['J16'].value = moved_to_tt_retenetion_fdr_issue
ws['J17'].value = business_accepted_defect_retenetion_fdr_issue+knownissue_retenetion_fdr_issue
ws['J18'].value = deployment_retenetion_fdr_issue
ws['J20'].value = cantreproduce_retenetion_fdr_issue
ws['J21'].value = duplicate_retenetion_fdr_issue
ws['J23'].value = fixed_retenetion_fdr_issue + pdfixed_retenetion_fdr_issue + postlaunch_fix_retenetion_fdr_issue
ws['J24'].value = resolved_retenetion_fdr_issue
ws['J25'].value = open_retenetion_fdr_issue
ws['J26'].value = blocker_retenetion_fdr_issue
ws['J27'].value = blocker_open_retenetion_fdr_issue
ws['J28'].value = blocker_resolved_retenetion_fdr_issue
ws['J34'].value = triaged_retenetion_fdr_issue
ws['J35'].value = valid_defect_ratio_retenetion_fdr
ws['J36'].value = defect_fix_ratio_retenetion_fdr
#closing the opened excel file after writing in it
wb.save(path_write_xl)

#calculating total DA_Seller Central issues
for i in range(0,len(df)):
    if seller_central_tag in df.iloc[i,-1]:
        seller_central_issues += 1
print('\n')
print("Total DA_Seller Central Issue : ",seller_central_issues)

#calculating realoved issues for DA_Seller Central
for i in range(0,len(df)):
    if (('Closed' in df.iloc[i,3] or 'Resolved' in df.iloc[i,3]) and seller_central_tag in df.iloc[i,-1]):
        resolved_seller_central_issue += 1
print('Resolves issues for DA_Seller Central : ',resolved_seller_central_issue)

#calculating open issues for DA_Seller Central issue
for i in range(1,len(df)):
    if 'Open' in df.iloc[i,3] and seller_central_tag in df.iloc[i,-1]:
        open_seller_central_issue += 1
print("Unresolved DA_Seller Central Issue : ",open_seller_central_issue)

#calculating fixed issues for DA_Seller Central vertical
fixed_seller_central_issue = calculating_issues(len(df), fixed_tag, seller_central_tag)
print("Fixed DA_Seller Central Issues : ",fixed_seller_central_issue)

#calculating cannot reproduce issues for DA_Seller Central Vertical
cantreproduce_seller_central_issue = calculating_issues(len(df), cantreproduce_tag, seller_central_tag)
print("Cannot Reproduce DA_Seller Central Issues : ",cantreproduce_seller_central_issue)

#calculating duplicate issues for DA_Seller Central Vertical
duplicate_seller_central_issue = calculating_issues(len(df), duplicate_tag, seller_central_tag)
print("Duplicate DA_Seller Central Issues : ",duplicate_seller_central_issue)

#calculating triaged issues for DA_Seller Central Vertical
triaged_seller_central_issue = calculating_issues(len(df), triaged_tag, seller_central_tag)
print("Triaged DA_Seller Central Issues : ",triaged_seller_central_issue)

#calcualting deferred issues for DA_Seller Central Vertical
deferred_seller_central_issue = calculating_issues(len(df), deferred_tag, seller_central_tag)
print("Deferred DA_Seller Central Issue : ",deferred_seller_central_issue)

#calculating rca_dispute issue for DA_Seller Central vertical
rcadispute_seller_central_issue = calculating_issues(len(df), rcadispute_tag, seller_central_tag)
print("RCA_Dispute DA_Seller Central Issue : ",rcadispute_seller_central_issue)

#calculating tc issue for DA_Seller Central vertical
tc_seller_central_issue = calculating_issues(len(df), tc_tag, seller_central_tag)
print("TC DA_Seller Central Issue : ",tc_seller_central_issue)

#calculating adhoc issue for DA_Seller Central vertical
adhoc_seller_central_issue = calculating_issues(len(df), adhoc_tag, seller_central_tag)
print("Adhoc DA_Seller Central Issue : ",adhoc_seller_central_issue)

#calculating moved_to_tt_issues for DA_Seller Central vertical
moved_to_tt_seller_central_issue = calculating_issues(len(df), moved_to_tt_tag, seller_central_tag)
print("Moved to TT for DA_Seller Central: ", moved_to_tt_seller_central_issue)

#calculating not_fixed_valid for DA_Seller Central vertical
notfixed_valid_seller_central_issue = calculating_issues(len(df), notfixed_valid_tag, seller_central_tag)
print("Not Fixed Valid DA_Seller Central Issue : ", notfixed_valid_seller_central_issue)

#calculating pdfixed for DA_Seller Central vertical
pdfixed_seller_central_issue = calculating_issues(len(df), pdfixed_tag, seller_central_tag)
print("PD_Fixed DA_Seller Central Issue : ", pdfixed_seller_central_issue)

#calculating postlaunch_fixed for DA_Seller Central vertical
postlaunch_fix_seller_central_issue = calculating_issues(len(df), postlaunch_fix_tag, seller_central_tag)
print("Post Launch Fix DA_Seller Central Issue : ", postlaunch_fix_seller_central_issue)

#calculating deployment issue for DA_Seller Central vertical
deployment_seller_central_issue = calculating_issues(len(df), deployment_tag, seller_central_tag)
print("Deployment DA_Seller Central Issue : ", deployment_seller_central_issue)

#calculating bydesign_valid for DA_Seller Central vertical
bydesign_valid_seller_central_issue = calculating_issues(len(df), bydesign_valid_tag, seller_central_tag)
print("By Design Valid DA_Seller Central Issue : ", bydesign_valid_seller_central_issue)

#calculating bydesign_tc for DA_Seller Central vertical
bydesign_tc_seller_central_issue = calculating_issues(len(df), bydesign_tc_tag, seller_central_tag)
print("ByDesign TC DA_Seller Central Issue : ", bydesign_tc_seller_central_issue)

#calculating business accepted defect for DA_Seller Central vertical
business_accepted_defect_seller_central_issue = calculating_issues(len(df), business_accepted_defect_tag, seller_central_tag)
print("Business Accepted Defect for DA_Seller Central : ", business_accepted_defect_seller_central_issue)

#calculating known issue for DA_Seller Central vertical
knownissue_seller_central_issue = calculating_issues(len(df), knownissue_tag, seller_central_tag)
print("Known Issue for  DA_Seller Central Vertical : ", knownissue_seller_central_issue)

#calulating blocker issue for DA_Seller Central vertical
blocker_seller_central_issue = calculating_blocker_issues(len(df), seller_central_tag)
print("DA_Seller Central Blocker Issues : ",blocker_seller_central_issue)

#calculating open and resolved blocker bugs for DA_Seller Central vertical
blocker_open_seller_central_issue, blocker_resolved_seller_central_issue = calcualting_open_and_resolved_blocker_issues(len(df), seller_central_tag)
print("DA_Seller Central Open Blocker Issues : ", blocker_open_seller_central_issue)
print("DA_Seller Central Resolved Blocker Issues : ", blocker_resolved_seller_central_issue)

#calculating valid issues for DA_Seller Central vertical
valid_seller_central_issue = (fixed_seller_central_issue+notfixed_valid_seller_central_issue+pdfixed_seller_central_issue+postlaunch_fix_seller_central_issue+ 
                              deferred_seller_central_issue+deployment_seller_central_issue+rcadispute_seller_central_issue+bydesign_valid_seller_central_issue+
                              bydesign_tc_seller_central_issue+business_accepted_defect_seller_central_issue)
print('Valid defects in DA_Seller Central vertical : ',valid_seller_central_issue)

#valid defect ration for DA_Seller Central Vertical
if resolved_seller_central_issue == 0:
    valid_defect_ratio_seller_central = 0
else:
    valid_defect_ratio_seller_central = int(valid_seller_central_issue*100/resolved_seller_central_issue)
print(str('Valid defect ration in DA_Seller Central : ')+str(valid_defect_ratio_seller_central)+'%')


#defect fix ratio for DA_Seller Central Vertical
if resolved_seller_central_issue == 0:
    defect_fix_ratio_seller_central = 0
else:
    defect_fix_ratio_seller_central = int(fixed_seller_central_issue*100/resolved_seller_central_issue)
print(str('Defect fix ration for DA_Seller Central : ')+str(defect_fix_ratio_seller_central)+'%')

#writing the values to MBR Format excel sheet for DA_Seller Central
wb = load_workbook(path_write_xl)
#selceting the active sheet from the excel file, active sheet is generally the sheet that we see when the excel file is opened
ws = wb.active
ws['L10'].value = seller_central_issues
ws['L11'].value = tc_seller_central_issue
ws['L12'].value = adhoc_seller_central_issue
ws['L13'].value = valid_seller_central_issue
ws['L14'].value = bydesign_tc_seller_central_issue+bydesign_valid_seller_central_issue
ws['L15'].value = rcadispute_acuistion_issue+deployment_seller_central_issue
ws['L16'].value = moved_to_tt_seller_central_issue
ws['L17'].value = business_accepted_defect_seller_central_issue+knownissue_seller_central_issue
ws['L18'].value = deployment_seller_central_issue
ws['L20'].value = cantreproduce_seller_central_issue
ws['L21'].value = duplicate_seller_central_issue
ws['L23'].value = fixed_seller_central_issue + pdfixed_seller_central_issue + postlaunch_fix_seller_central_issue
ws['L24'].value = resolved_seller_central_issue
ws['L25'].value = open_seller_central_issue
ws['L26'].value = blocker_seller_central_issue
ws['L27'].value = blocker_open_seller_central_issue
ws['L28'].value = blocker_resolved_seller_central_issue
ws['L34'].value = triaged_seller_central_issue
ws['L35'].value = valid_defect_ratio_seller_central
ws['L36'].value = defect_fix_ratio_seller_central
#closing the opened excel file after writing in it
wb.save(path_write_xl)

#calculating total DA_Benefits issues
for i in range(0,len(df)):
    if benefits_tag in df.iloc[i,-1]:
        benefits_issues += 1
print('\n')
print("Total DA_Benefits Issue : ",benefits_issues)

#calculating realoved issues for DA_Benefits
for i in range(0,len(df)):
    if (('Closed' in df.iloc[i,3] or 'Resolved' in df.iloc[i,3]) and benefits_tag in df.iloc[i,-1]):
        resolved_benefits_issue += 1
print('Resolves issues for DA_Benefits : ',resolved_benefits_issue)

#calculating open issues for DA_Benefits issue
for i in range(1,len(df)):
    if 'Open' in df.iloc[i,3] and benefits_tag in df.iloc[i,-1]:
        open_benefits_issue += 1
print("Unresolved DA_Benefits Issue : ",open_benefits_issue)

#calculating fixed issues for DA_Benefits vertical
fixed_benefits_issue = calculating_issues(len(df), fixed_tag, benefits_tag)
print("Fixed DA_Benefits Issues : ",fixed_benefits_issue)

#calculating cannot reproduce issues for DA_Benefits Vertical
cantreproduce_benefits_issue = calculating_issues(len(df), cantreproduce_tag, benefits_tag)
print("Cannot Reproduce DA_Benefits Issues : ",cantreproduce_benefits_issue)

#calculating duplicate issues for DA_Benefits Vertical
duplicate_benefits_issue = calculating_issues(len(df), duplicate_tag, benefits_tag)
print("Duplicate DA_Benefits Issues : ",duplicate_benefits_issue)

#calculating triaged issues for DA_Benefits Vertical
triaged_benefits_issue = calculating_issues(len(df), triaged_tag, benefits_tag)
print("Triaged DA_Benefits Issues : ",triaged_benefits_issue)

#calcualting deferred issues for DA_Benefits Vertical
deferred_benefits_issue = calculating_issues(len(df), deferred_tag, benefits_tag)
print("Deferred DA_Benefits Issue : ",deferred_benefits_issue)

#calculating rca_dispute issue for DA_Benefits vertical
rcadispute_benefits_issue = calculating_issues(len(df), rcadispute_tag, benefits_tag)
print("RCA_Dispute DA_Benefits Issue : ",rcadispute_benefits_issue)

#calculating tc issue for DA_Benefits vertical
tc_benefits_issue = calculating_issues(len(df), tc_tag, benefits_tag)
print("TC DA_Benefits Issue : ",tc_benefits_issue)

#calculating adhoc issue for DA_Benefits vertical
adhoc_benefits_issue = calculating_issues(len(df), adhoc_tag, benefits_tag)
print("Adhoc DA_Benefits Issue : ",adhoc_benefits_issue)

#calculating moved_to_tt_issues for DA_Benefits vertical
moved_to_tt_benefits_issue = calculating_issues(len(df), moved_to_tt_tag, benefits_tag)
print("Moved to TT for DA_Benefits: ", moved_to_tt_benefits_issue)

#calculating not_fixed_valid for DA_Benefits vertical
notfixed_valid_benefits_issue = calculating_issues(len(df), notfixed_valid_tag, benefits_tag)
print("Not Fixed Valid DA_Benefits Issue : ", notfixed_valid_benefits_issue)

#calculating pdfixed for DA_Benefits vertical
pdfixed_benefits_issue = calculating_issues(len(df), pdfixed_tag, benefits_tag)
print("PD_Fixed DA_Benefits Issue : ", pdfixed_benefits_issue)

#calculating postlaunch_fixed for DA_Benefits vertical
postlaunch_fix_benefits_issue = calculating_issues(len(df), postlaunch_fix_tag, benefits_tag)
print("Post Launch Fix DA_Benefits Issue : ", postlaunch_fix_benefits_issue)

#calculating deployment issue for DA_Benefits vertical
deployment_benefits_issue = calculating_issues(len(df), deployment_tag, benefits_tag)
print("Deployment DA_Benefits Issue : ", deployment_benefits_issue)

#calculating bydesign_valid for DA_Benefits vertical
bydesign_valid_benefits_issue = calculating_issues(len(df), bydesign_valid_tag, benefits_tag)
print("By Design Valid DA_Benefits Issue : ", bydesign_valid_benefits_issue)

#calculating bydesign_tc for DA_Benefits vertical
bydesign_tc_benefits_issue = calculating_issues(len(df), bydesign_tc_tag, benefits_tag)
print("ByDesign TC DA_Benefits Issue : ", bydesign_tc_benefits_issue)

#calculating business accepted defect for DA_Benefits vertical
business_accepted_defect_benefits_issue = calculating_issues(len(df), business_accepted_defect_tag, benefits_tag)
print("Business Accepted Defect for DA_Benefits : ", business_accepted_defect_benefits_issue)

#calculating known issue for DA_Benefits vertical
knownissue_benefits_issue = calculating_issues(len(df), knownissue_tag, benefits_tag)
print("Known Issue for  DA_Benefits Vertical : ", knownissue_benefits_issue)

#calulating blocker issue for DA_Benefits vertical
blocker_benefits_issue = calculating_blocker_issues(len(df), benefits_tag)
print("DA_Benefits Blocker Issues : ",blocker_benefits_issue)

#calculating open and resolved blocker bugs for DA_Benefits vertical
blocker_open_benefits_issue, blocker_resolved_benefits_issue = calcualting_open_and_resolved_blocker_issues(len(df), benefits_tag)
print("DA_Benefits Open Blocker Issues : ", blocker_open_benefits_issue)
print("DA_Benefits Resolved Blocker Issues : ", blocker_resolved_benefits_issue)

#calculating valid issues for DA_Benefits vertical
valid_benefits_issue = (fixed_benefits_issue+notfixed_valid_benefits_issue+pdfixed_benefits_issue+postlaunch_fix_benefits_issue+ 
                        deferred_benefits_issue+deployment_benefits_issue+rcadispute_benefits_issue+bydesign_valid_benefits_issue+
                        bydesign_tc_benefits_issue+business_accepted_defect_benefits_issue)
print('Valid defects in DA_Benefits vertical : ',valid_benefits_issue)

#valid defect ration for DA_Benefits Vertical
if resolved_benefits_issue == 0:
    valid_defect_ratio_benefits = 0
else:
    valid_defect_ratio_benefits = int(valid_benefits_issue*100/resolved_benefits_issue)
print(str('Valid defect ration in DA_Benefits : ')+str(valid_defect_ratio_benefits)+'%')


#defect fix ratio for DA_Benefits Vertical
if resolved_benefits_issue == 0:
    defect_fix_ratio_benefits = 0
else:
    defect_fix_ratio_benefits = int(fixed_benefits_issue*100/resolved_benefits_issue)
print(str('Defect fix ration for DA_Benefits : ')+str(defect_fix_ratio_benefits)+'%')

#writing the values to MBR Format excel sheet for DA_Benefits
#TO-DO Figure out under which vertical the data for DA_Benefits to be written

#calculating total DA_PrimeCoreService issues
for i in range(0,len(df)):
    if prime_core_services_tag in df.iloc[i,-1]:
        prime_core_services_issues += 1
print('\n')
print("Total DA_PrimeCoreService Issue : ",prime_core_services_issues)

#calculating realoved issues for DA_PrimeCoreService
for i in range(0,len(df)):
    if (('Closed' in df.iloc[i,3] or 'Resolved' in df.iloc[i,3]) and prime_core_services_tag in df.iloc[i,-1]):
        resolved_prime_core_services_issue += 1
print('Resolves issues for DA_PrimeCoreService : ',resolved_prime_core_services_issue)

#calculating open issues for DA_PrimeCoreServiceissue
for i in range(1,len(df)):
    if 'Open' in df.iloc[i,3] and prime_core_services_tag in df.iloc[i,-1]:
        open_prime_core_services_issue += 1
print("Unresolved DA_PrimeCoreServiceIssue : ",open_prime_core_services_issue)

#calculating fixed issues for DA_PrimeCoreService vertical
fixed_prime_core_services_issue = calculating_issues(len(df), fixed_tag, prime_core_services_tag)
print("Fixed DA_PrimeCoreService Issues : ",fixed_prime_core_services_issue)

#calculating cannot reproduce issues for DA_PrimeCoreService Vertical
cantreproduce_prime_core_services_issue = calculating_issues(len(df), cantreproduce_tag, prime_core_services_tag)
print("Cannot Reproduce DA_PrimeCoreService Issues : ",cantreproduce_prime_core_services_issue)

#calculating duplicate issues for DA_PrimeCoreService Vertical
duplicate_prime_core_services_issue = calculating_issues(len(df), duplicate_tag, prime_core_services_tag)
print("Duplicate DA_PrimeCoreService Issues : ",duplicate_prime_core_services_issue)

#calculating triaged issues for DA_PrimeCoreService Vertical
triaged_prime_core_services_issue = calculating_issues(len(df), triaged_tag, prime_core_services_tag)
print("Triaged DA_PrimeCoreService Issues : ",triaged_prime_core_services_issue)

#calcualting deferred issues for DA_PrimeCoreService Vertical
deferred_prime_core_services_issue = calculating_issues(len(df), deferred_tag, prime_core_services_tag)
print("Deferred DA_PrimeCoreService Issue : ",deferred_prime_core_services_issue)

#calculating rca_dispute issue for DA_PrimeCoreService vertical
rcadispute_prime_core_services_issue = calculating_issues(len(df), rcadispute_tag, prime_core_services_tag)
print("RCA_Dispute DA_PrimeCoreService Issue : ",rcadispute_prime_core_services_issue)

#calculating tc issue for DA_PrimeCoreService vertical
tc_prime_core_services_issue = calculating_issues(len(df), tc_tag, prime_core_services_tag)
print("TC DA_PrimeCoreService Issue : ",tc_prime_core_services_issue)

#calculating adhoc issue for DA_PrimeCoreService vertical
adhoc_prime_core_services_issue = calculating_issues(len(df), adhoc_tag, prime_core_services_tag)
print("Adhoc DA_PrimeCoreService Issue : ",adhoc_prime_core_services_issue)

#calculating moved_to_tt_issues for DA_PrimeCoreService vertical
moved_to_tt_prime_core_services_issue = calculating_issues(len(df), moved_to_tt_tag, prime_core_services_tag)
print("Moved to TT for DA_PrimeCoreService: ", moved_to_tt_prime_core_services_issue)

#calculating not_fixed_valid for DA_PrimeCoreService vertical
notfixed_valid_prime_core_services_issue = calculating_issues(len(df), notfixed_valid_tag, prime_core_services_tag)
print("Not Fixed Valid DA_PrimeCoreService Issue : ", notfixed_valid_prime_core_services_issue)

#calculating pdfixed for DA_PrimeCoreService vertical
pdfixed_prime_core_services_issue = calculating_issues(len(df), pdfixed_tag, prime_core_services_tag)
print("PD_Fixed DA_PrimeCoreService Issue : ", pdfixed_prime_core_services_issue)

#calculating postlaunch_fixed for DA_PrimeCoreService vertical
postlaunch_fix_prime_core_services_issue = calculating_issues(len(df), postlaunch_fix_tag, prime_core_services_tag)
print("Post Launch Fix DA_PrimeCoreService Issue : ", postlaunch_fix_prime_core_services_issue)

#calculating deployment issue for DA_PrimeCoreService vertical
deployment_prime_core_services_issue = calculating_issues(len(df), deployment_tag, prime_core_services_tag)
print("Deployment DA_PrimeCoreService Issue : ", deployment_prime_core_services_issue)

#calculating bydesign_valid for DA_PrimeCoreService vertical
bydesign_valid_prime_core_services_issue = calculating_issues(len(df), bydesign_valid_tag, prime_core_services_tag)
print("By Design Valid DA_PrimeCoreService Issue : ", bydesign_valid_prime_core_services_issue)

#calculating bydesign_tc for DA_PrimeCoreService vertical
bydesign_tc_prime_core_services_issue = calculating_issues(len(df), bydesign_tc_tag, prime_core_services_tag)
print("ByDesign TC DA_PrimeCoreService Issue : ", bydesign_tc_prime_core_services_issue)

#calculating business accepted defect for DA_PrimeCoreService vertical
business_accepted_defect_prime_core_services_issue = calculating_issues(len(df), business_accepted_defect_tag, prime_core_services_tag)
print("Business Accepted Defect for DA_PrimeCoreService : ", business_accepted_defect_prime_core_services_issue)

#calculating known issue for DA_PrimeCoreService vertical
knownissue_prime_core_services_issue = calculating_issues(len(df), knownissue_tag, prime_core_services_tag)
print("Known Issue for  DA_PrimeCoreService Vertical : ", knownissue_prime_core_services_issue)

#calulating blocker issue for DA_PrimeCoreService vertical
blocker_prime_core_services_issue = calculating_blocker_issues(len(df), prime_core_services_tag)
print("DA_PrimeCoreService Blocker Issues : ",blocker_prime_core_services_issue)

#calculating open and resolved blocker bugs for DA_PrimeCoreService vertical
blocker_open_prime_core_services_issue, blocker_resolved_prime_core_services_issue = calcualting_open_and_resolved_blocker_issues(len(df), prime_core_services_tag)
print("DA_PrimeCoreService Open Blocker Issues : ", blocker_open_prime_core_services_issue)
print("DA_PrimeCoreService Resolved Blocker Issues : ", blocker_resolved_prime_core_services_issue)


#calculating valid issues for DA_PrimeCoreService vertical
valid_prime_core_services_issue = (fixed_prime_core_services_issue+notfixed_valid_prime_core_services_issue+pdfixed_prime_core_services_issue+postlaunch_fix_prime_core_services_issue+ 
                    deferred_prime_core_services_issue+deployment_prime_core_services_issue+rcadispute_prime_core_services_issue+bydesign_valid_prime_core_services_issue+
                    bydesign_tc_prime_core_services_issue+business_accepted_defect_prime_core_services_issue)
print('Valid defects in DA_PrimeCoreService vertical : ',valid_prime_core_services_issue)

#valid defect ration for DA_PrimeCoreService Vertical
if resolved_prime_core_services_issue == 0:
    valid_defect_ratio_prime_core_services = 0
else:
    valid_defect_ratio_prime_core_services = int(valid_prime_core_services_issue*100/resolved_prime_core_services_issue)
print(str('Valid defect ration in DA_PrimeCoreService : ')+str(valid_defect_ratio_prime_core_services)+'%')


#defect fix ratio for DA_PrimeCoreService Vertical
if resolved_prime_core_services_issue == 0:
    defect_fix_ratio_prime_core_services = 0
else:
    defect_fix_ratio_prime_core_services = int(fixed_prime_core_services_issue*100/resolved_prime_core_services_issue)
print(str('Defect fix ration for DA_PrimeCoreService Central : ')+str(defect_fix_ratio_prime_core_services)+'%')

#writing the values to MBR Format excel sheet for DA_PrimeCoreService Vertical
wb = load_workbook(path_write_xl)
#selceting the active sheet from the excel file, active sheet is generally the sheet that we see when the excel file is opened
ws = wb.active
ws['R10'].value = prime_core_services_issues
ws['R11'].value = tc_prime_core_services_issue
ws['R12'].value = adhoc_prime_core_services_issue
ws['R13'].value = valid_prime_core_services_issue
ws['R14'].value = bydesign_tc_prime_core_services_issue+bydesign_valid_prime_core_services_issue
ws['R15'].value = rcadispute_acuistion_issue+deployment_prime_core_services_issue
ws['R16'].value = moved_to_tt_prime_core_services_issue
ws['R17'].value = business_accepted_defect_prime_core_services_issue+knownissue_prime_core_services_issue
ws['R18'].value = deployment_prime_core_services_issue
ws['R20'].value = cantreproduce_prime_core_services_issue
ws['R21'].value = duplicate_prime_core_services_issue
ws['R23'].value = fixed_prime_core_services_issue + pdfixed_prime_core_services_issue + postlaunch_fix_prime_core_services_issue
ws['R24'].value = resolved_prime_core_services_issue
ws['R25'].value = open_prime_core_services_issue
ws['R26'].value = blocker_prime_core_services_issue
ws['R27'].value = blocker_open_prime_core_services_issue
ws['R28'].value = blocker_resolved_prime_core_services_issue
ws['R34'].value = triaged_prime_core_services_issue
ws['R35'].value = valid_defect_ratio_prime_core_services
ws['R36'].value = defect_fix_ratio_prime_core_services
#closing the opened excel file after writing in it
wb.save(path_write_xl)

#calculating total DA_Activation&Membership issues
for i in range(0,len(df)):
    if activation_and_membership_tag in df.iloc[i,-1]:
        activation_and_membership_issues += 1
print('\n')
print("Total DA_Activation&Membership Issue : ",activation_and_membership_issues)

#calculating realoved issues for DA_Activation&Membership
for i in range(0,len(df)):
    if (('Closed' in df.iloc[i,3] or 'Resolved' in df.iloc[i,3]) and activation_and_membership_tag in df.iloc[i,-1]):
        resolved_activation_and_membership_issue += 1
print('Resolves issues for DA_Activation&Membership : ',resolved_activation_and_membership_issue)

#calculating open issues for DA_Activation&Membership issue
for i in range(1,len(df)):
    if 'Open' in df.iloc[i,3] and activation_and_membership_tag in df.iloc[i,-1]:
        open_activation_and_membership_issue += 1
print("Unresolved DA_Activation&Membership Issue : ",open_activation_and_membership_issue)

#calculating fixed issues for DA_Activation&Membership vertical
fixed_activation_and_membership_issue = calculating_issues(len(df), fixed_tag, activation_and_membership_tag)
print("Fixed DA_Activation&Membership Issues : ",fixed_activation_and_membership_issue)

#calculating cannot reproduce issues for DA_Activation&Membership Vertical
cantreproduce_activation_and_membership_issue = calculating_issues(len(df), cantreproduce_tag, activation_and_membership_tag)
print("Cannot Reproduce DA_Activation&Membership Issues : ",cantreproduce_activation_and_membership_issue)

#calculating duplicate issues for DA_Activation&Membership Vertical
duplicate_activation_and_membership_issue = calculating_issues(len(df), duplicate_tag, activation_and_membership_tag)
print("Duplicate DA_Activation&Membership Issues : ",duplicate_activation_and_membership_issue)

#calculating triaged issues for DA_Activation&Membership Vertical
triaged_activation_and_membership_issue = calculating_issues(len(df), triaged_tag, activation_and_membership_tag)
print("Triaged DA_Activation&Membership Issues : ",triaged_activation_and_membership_issue)

#calcualting deferred issues for DA_Activation&Membership Vertical
deferred_activation_and_membership_issue = calculating_issues(len(df), deferred_tag, activation_and_membership_tag)
print("Deferred DA_Activation&Membership Issue : ",deferred_activation_and_membership_issue)

#calculating rca_dispute issue for DA_Activation&Membership vertical
rcadispute_activation_and_membership_issue = calculating_issues(len(df), rcadispute_tag, activation_and_membership_tag)
print("RCA_Dispute DA_Activation&Membership Issue : ",rcadispute_activation_and_membership_issue)

#calculating tc issue for DA_Activation&Membership vertical
tc_activation_and_membership_issue = calculating_issues(len(df), tc_tag, activation_and_membership_tag)
print("TC DA_Activation&Membership Issue : ",tc_activation_and_membership_issue)

#calculating adhoc issue for DA_Activation&Membership vertical
adhoc_activation_and_membership_issue = calculating_issues(len(df), adhoc_tag, activation_and_membership_tag)
print("Adhoc DA_Activation&Membership Issue : ",adhoc_activation_and_membership_issue)

#calculating moved_to_tt_issues for DA_Activation&Membership vertical
moved_to_tt_activation_and_membership_issue = calculating_issues(len(df), moved_to_tt_tag, activation_and_membership_tag)
print("Moved to TT for DA_Activation&Membership: ", moved_to_tt_activation_and_membership_issue)

#calculating not_fixed_valid for DA_Activation&Membership vertical
notfixed_valid_activation_and_membership_issue = calculating_issues(len(df), notfixed_valid_tag, activation_and_membership_tag)
print("Not Fixed Valid DA_Activation&Membership Issue : ", notfixed_valid_activation_and_membership_issue)

#calculating pdfixed for DA_Activation&Membership vertical
pdfixed_activation_and_membership_issue = calculating_issues(len(df), pdfixed_tag, activation_and_membership_tag)
print("PD_Fixed DA_Activation&Membership Issue : ", pdfixed_activation_and_membership_issue)

#calculating postlaunch_fixed for DA_Activation&Membership vertical
postlaunch_fix_activation_and_membership_issue = calculating_issues(len(df), postlaunch_fix_tag, activation_and_membership_tag)
print("Post Launch Fix DA_Activation&Membership Issue : ", postlaunch_fix_activation_and_membership_issue)

#calculating deployment issue for DA_Activation&Membership vertical
deployment_activation_and_membership_issue = calculating_issues(len(df), deployment_tag, activation_and_membership_tag)
print("Deployment DA_Activation&Membership Issue : ", deployment_activation_and_membership_issue)

#calculating bydesign_valid for DA_Activation&Membership vertical
bydesign_valid_activation_and_membership_issue = calculating_issues(len(df), bydesign_valid_tag, activation_and_membership_tag)
print("By Design Valid DA_Activation&Membership Issue : ", bydesign_valid_activation_and_membership_issue)

#calculating bydesign_tc for DA_Activation&Membership vertical
bydesign_tc_activation_and_membership_issue = calculating_issues(len(df), bydesign_tc_tag, activation_and_membership_tag)
print("ByDesign TC DA_Activation&Membership Issue : ", bydesign_tc_activation_and_membership_issue)

#calculating business accepted defect for DA_Activation&Membership vertical
business_accepted_defect_activation_and_membership_issue = calculating_issues(len(df), business_accepted_defect_tag, activation_and_membership_tag)
print("Business Accepted Defect for DA_Activation&Membership : ", business_accepted_defect_activation_and_membership_issue)

#calculating known issue for DA_Activation&Membership vertical
knownissue_activation_and_membership_issue = calculating_issues(len(df), knownissue_tag, activation_and_membership_tag)
print("Known Issue for  DA_Activation&Membership Vertical : ", knownissue_activation_and_membership_issue)

#calulating blocker issue for DA_Activation&Membership vertical
blocker_activation_and_membership_issue = calculating_blocker_issues(len(df), activation_and_membership_tag)
print("DA_Activation&Membership Blocker Issues : ",blocker_activation_and_membership_issue)

#calculating open and resolved blocker bugs for DA_Activation&Membership vertical
blocker_open_activation_and_membership_issue, blocker_resolved_activation_and_membership_issue = calcualting_open_and_resolved_blocker_issues(len(df), activation_and_membership_tag)
print("DA_Activation&Membership Open Blocker Issues : ", blocker_open_activation_and_membership_issue)
print("DA_Activation&Membership Resolved Blocker Issues : ", blocker_resolved_activation_and_membership_issue)

#calculating valid issues for DA_Activation&Membership vertical
valid_activation_and_membership_issue = (fixed_activation_and_membership_issue+notfixed_valid_activation_and_membership_issue+pdfixed_activation_and_membership_issue+postlaunch_fix_activation_and_membership_issue+ 
                    deferred_activation_and_membership_issue+deployment_activation_and_membership_issue+rcadispute_activation_and_membership_issue+bydesign_valid_activation_and_membership_issue+
                    bydesign_tc_activation_and_membership_issue+business_accepted_defect_activation_and_membership_issue)
print('Valid defects in DA_Activation&Membership vertical : ',valid_activation_and_membership_issue)

#valid defect ration for DA_Activation&Membership Vertical
if resolved_activation_and_membership_issue == 0:
    valid_defect_ratio_activation_and_membership = 0
else:
    valid_defect_ratio_activation_and_membership = int(valid_activation_and_membership_issue*100/resolved_activation_and_membership_issue)
print(str('Valid defect ration in DA_Activation&Membership : ')+str(valid_defect_ratio_activation_and_membership)+'%')


#defect fix ratio for DA_Activation&Membership Vertical
if resolved_activation_and_membership_issue == 0:
    defect_fix_ratio_activation_and_membership = 0
else:
    defect_fix_ratio_activation_and_membership = int(fixed_activation_and_membership_issue*100/resolved_activation_and_membership_issue)
print(str('Defect fix ration for DA_Activation&Membership Central : ')+str(defect_fix_ratio_activation_and_membership)+'%')

#writing the values to MBR Format excel sheet for DA_Activation&Membership Vertical

#calculating total DA_Retention_Benefits_Propositions issues
for i in range(0,len(df)):
    if retention_benefits_propositions_tag in df.iloc[i,-1]:
        retention_benefits_propositions_issues += 1
print('\n')
print("Total DA_Retention_Benefits_Propositions Issue : ",retention_benefits_propositions_issues)

#calculating realoved issues for DA_Retention_Benefits_Propositions
for i in range(0,len(df)):
    if (('Closed' in df.iloc[i,3] or 'Resolved' in df.iloc[i,3]) and retention_benefits_propositions_tag in df.iloc[i,-1]):
        resolved_retention_benefits_propositions_issue += 1
print('Resolves issues for DA_Retention_Benefits_Propositions : ',resolved_retention_benefits_propositions_issue)

#calculating open issues for DA_Retention_Benefits_Propositions issue
for i in range(1,len(df)):
    if 'Open' in df.iloc[i,3] and retention_benefits_propositions_tag in df.iloc[i,-1]:
        open_retention_benefits_propositions_issue += 1
print("Unresolved DA_Retention_Benefits_Propositions Issue : ",open_retention_benefits_propositions_issue)

#calculating fixed issues for DA_Retention_Benefits_Propositions vertical
fixed_retention_benefits_propositions_issue = calculating_issues(len(df), fixed_tag, retention_benefits_propositions_tag)
print("Fixed DA_Retention_Benefits_Propositions Issues : ",fixed_retention_benefits_propositions_issue)

#calculating cannot reproduce issues for DA_Retention_Benefits_Propositions Vertical
cantreproduce_retention_benefits_propositions_issue = calculating_issues(len(df), cantreproduce_tag, retention_benefits_propositions_tag)
print("Cannot Reproduce DA_Retention_Benefits_Propositions Issues : ",cantreproduce_retention_benefits_propositions_issue)

#calculating duplicate issues for DA_Retention_Benefits_Propositions Vertical
duplicate_retention_benefits_propositions_issue = calculating_issues(len(df), duplicate_tag, retention_benefits_propositions_tag)
print("Duplicate DA_Retention_Benefits_Propositions Issues : ",duplicate_retention_benefits_propositions_issue)

#calculating triaged issues for DA_Retention_Benefits_Propositions Vertical
triaged_retention_benefits_propositions_issue = calculating_issues(len(df), triaged_tag, retention_benefits_propositions_tag)
print("Triaged DA_Retention_Benefits_Propositions Issues : ",triaged_retention_benefits_propositions_issue)

#calcualting deferred issues for DA_Retention_Benefits_Propositions Vertical
deferred_retention_benefits_propositions_issue = calculating_issues(len(df), deferred_tag, retention_benefits_propositions_tag)
print("Deferred DA_Retention_Benefits_Propositions Issue : ",deferred_retention_benefits_propositions_issue)

#calculating rca_dispute issue for DA_Retention_Benefits_Propositions vertical
rcadispute_retention_benefits_propositions_issue = calculating_issues(len(df), rcadispute_tag, retention_benefits_propositions_tag)
print("RCA_Dispute DA_Retention_Benefits_Propositions Issue : ",rcadispute_retention_benefits_propositions_issue)

#calculating tc issue for DA_Retention_Benefits_Propositions vertical
tc_retention_benefits_propositions_issue = calculating_issues(len(df), tc_tag, retention_benefits_propositions_tag)
print("TC DA_Retention_Benefits_Propositions Issue : ",tc_retention_benefits_propositions_issue)

#calculating adhoc issue for DA_Retention_Benefits_Propositions vertical
adhoc_retention_benefits_propositions_issue = calculating_issues(len(df), adhoc_tag, retention_benefits_propositions_tag)
print("Adhoc DA_Retention_Benefits_Propositions Issue : ",adhoc_retention_benefits_propositions_issue)

#calculating moved_to_tt_issues for DA_Retention_Benefits_Propositions vertical
moved_to_tt_retention_benefits_propositions_issue = calculating_issues(len(df), moved_to_tt_tag, retention_benefits_propositions_tag)
print("Moved to TT for DA_Retention_Benefits_Propositions: ", moved_to_tt_retention_benefits_propositions_issue)

#calculating not_fixed_valid for DA_Retention_Benefits_Propositions vertical
notfixed_valid_retention_benefits_propositions_issue = calculating_issues(len(df), notfixed_valid_tag, retention_benefits_propositions_tag)
print("Not Fixed Valid DA_Retention_Benefits_Propositions Issue : ", notfixed_valid_retention_benefits_propositions_issue)

#calculating pdfixed for DA_Retention_Benefits_Propositions vertical
pdfixed_retention_benefits_propositions_issue = calculating_issues(len(df), pdfixed_tag, retention_benefits_propositions_tag)
print("PD_Fixed DA_Retention_Benefits_Propositions Issue : ", pdfixed_retention_benefits_propositions_issue)

#calculating postlaunch_fixed for DA_Retention_Benefits_Propositions vertical
postlaunch_fix_retention_benefits_propositions_issue = calculating_issues(len(df), postlaunch_fix_tag, retention_benefits_propositions_tag)
print("Post Launch Fix DA_Retention_Benefits_Propositions Issue : ", postlaunch_fix_retention_benefits_propositions_issue)

#calculating deployment issue for DA_Retention_Benefits_Propositions vertical
deployment_retention_benefits_propositions_issue = calculating_issues(len(df), deployment_tag, retention_benefits_propositions_tag)
print("Deployment DA_Retention_Benefits_Propositions Issue : ", deployment_retention_benefits_propositions_issue)

#calculating bydesign_valid for DA_Retention_Benefits_Propositions vertical
bydesign_valid_retention_benefits_propositions_issue = calculating_issues(len(df), bydesign_valid_tag, retention_benefits_propositions_tag)
print("By Design Valid DA_Retention_Benefits_Propositions Issue : ", bydesign_valid_retention_benefits_propositions_issue)

#calculating bydesign_tc for DA_Retention_Benefits_Propositions vertical
bydesign_tc_retention_benefits_propositions_issue = calculating_issues(len(df), bydesign_tc_tag, retention_benefits_propositions_tag)
print("ByDesign TC DA_Retention_Benefits_Propositions Issue : ", bydesign_tc_retention_benefits_propositions_issue)

#calculating business accepted defect for DA_Retention_Benefits_Propositions vertical
business_accepted_defect_retention_benefits_propositions_issue = calculating_issues(len(df), business_accepted_defect_tag, retention_benefits_propositions_tag)
print("Business Accepted Defect for DA_Retention_Benefits_Propositions : ", business_accepted_defect_retention_benefits_propositions_issue)

#calculating known issue for DA_Retention_Benefits_Propositions vertical
knownissue_retention_benefits_propositions_issue = calculating_issues(len(df), knownissue_tag, retention_benefits_propositions_tag)
print("Known Issue for  DA_Retention_Benefits_Propositions Vertical : ", knownissue_retention_benefits_propositions_issue)

#calulating blocker issue for DA_Retention_Benefits_Propositions vertical
blocker_issue_retention_benefits_propositions = calculating_blocker_issues(len(df), retention_benefits_propositions_tag)
print("DA_Retention_Benefits_Propositions Blocker Issues : ",blocker_issue_retention_benefits_propositions)

#calculating open and resolved blocker bugs for DA_Retention_Benefits_Propositions vertical
blocker_open_issue_retention_benefits_propositions, blocker_resolved_issue_retention_benefits_propositions = calcualting_open_and_resolved_blocker_issues(len(df), retention_benefits_propositions_tag)
print("DA_Retention_Benefits_Propositions Open Blocker Issues : ", blocker_open_issue_retention_benefits_propositions)
print("DA_Retention_Benefits_Propositions Resolved Blocker Issues : ", blocker_resolved_issue_retention_benefits_propositions) 
    

#calculating valid issues for DA_Retention_Benefits_Propositions vertical
valid_retention_benefits_propositions_issue = (fixed_retention_benefits_propositions_issue+notfixed_valid_retention_benefits_propositions_issue+pdfixed_retention_benefits_propositions_issue+postlaunch_fix_retention_benefits_propositions_issue+ 
                    deferred_retention_benefits_propositions_issue+deployment_retention_benefits_propositions_issue+rcadispute_retention_benefits_propositions_issue+bydesign_valid_retention_benefits_propositions_issue+
                    bydesign_tc_retention_benefits_propositions_issue+business_accepted_defect_retention_benefits_propositions_issue)
print('Valid defects in DA_Retention_Benefits_Propositions vertical : ',valid_retention_benefits_propositions_issue)

#valid defect ration for DA_Retention_Benefits_Propositions Vertical
if resolved_retention_benefits_propositions_issue == 0:
    valid_defect_ratio_retention_benefits_propositions = 0
else:
    valid_defect_ratio_retention_benefits_propositions = int(valid_retention_benefits_propositions_issue*100/resolved_retention_benefits_propositions_issue)
print(str('Valid defect ration in DA_Retention_Benefits_Propositions : ')+str(valid_defect_ratio_retention_benefits_propositions)+'%')


#defect fix ratio for DA_Retention_Benefits_Propositions Vertical
if resolved_retention_benefits_propositions_issue == 0:
    defect_fix_ratio_retention_benefits_propositions = 0
else:
    defect_fix_ratio_retention_benefits_propositions = int(fixed_retention_benefits_propositions_issue*100/resolved_retention_benefits_propositions_issue)
print(str('Defect fix ration for DA_Retention_Benefits_Propositions Central : ')+str(defect_fix_ratio_retention_benefits_propositions)+'%')

#writing the values to MBR Format excel sheet for DA_Retention_Benefits_Propositions Vertical

#writing the values to the "Total" column of the MBR Format excel sheet
def calculating_values_for_total_column(row_value):
    cell_range = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M' , 'N', 'O' , 'P', 'Q', 'R']
    total_sum = 0
    wb = load_workbook(path_write_xl)
    ws = wb.active
    for cell_name in cell_range:
        if ws[cell_name+row_value].value == None:
            total_sum += 0
        else:
            total_sum += ws[cell_name+row_value].value
    wb.save(path_write_xl)
    return total_sum

#writing to the Total column of the MBR Format excel sheet    
wb = load_workbook(path_write_xl)
#selceting the active sheet from the excel file, active sheet is generally the sheet that we see when the excel file is opened
ws = wb.active

#writing total count of bugs identified to MBR Format excel sheet total column
ws['S10'] = calculating_values_for_total_column('10')

#writing total count of tc bugs to MBR Format excel sheet total column 
ws['S11'] = calculating_values_for_total_column('11')

#writing total count of adhoc bugs to MBR Format excel sheet total column
ws['S12'] = calculating_values_for_total_column('12')

#writing total count of valid bugs to MBR Format excel sheet total column
ws['S13'] = calculating_values_for_total_column('13')

#writing total count of By Design issue to MBR Format excel sheet total column
ws['S14'] = calculating_values_for_total_column('14')

#writing total count of Deployment/RCA Dispute bugs to MBR Format excel sheet total column
ws['S15'] = calculating_values_for_total_column('15')

#writing total count of Moved to TT bugs to MBR Format excel sheet total column
ws['S16'] = calculating_values_for_total_column('16')

#writing total count of Deferred/Known Issues/Business Accepted Defects to MBR Format excel sheet total column
ws['S17'] = calculating_values_for_total_column('17')

#writing total count of Weblab Issues to MBR Format excel sheet total column
ws['S18'] = calculating_values_for_total_column('18')

#writing total count of cannot reproduce issues to MBR Format excel sheet total column
ws['S20'] = calculating_values_for_total_column('20')

#writing total count of duplicate issues to MBR Format excel sheet total column
ws['S21'] = calculating_values_for_total_column('21')

#writing total count of fixed bug to MBR Format excel sheet total column
ws['S23'] = calculating_values_for_total_column('23')

#writing total count of resolved bugs to MBR Format excel sheet total column
ws['S24'] = calculating_values_for_total_column('24')

#writing total count of open bugs to MBR Format excel sheet total column
ws['S25'] = calculating_values_for_total_column('25')

#writing total count of triaged bugs to MBR Format excel sheet total coulmn
ws['S34'] = calculating_values_for_total_column('34')

#saving the excel sheet after writing in it
wb.save(path_write_xl)